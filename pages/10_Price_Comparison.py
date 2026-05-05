from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any
import re
import uuid

import pandas as pd
import streamlit as st

from core.auth import require_login
from database.connection import execute, get_connection
from services.upgrade_service import list_module_records, upsert_module_record
from ui.theme import apply_theme, render_page_header
from ui.upgrade_ui import render_upgrade_css, render_upgrade_intro, render_metric_grid, render_layered_records
from utils.dates import now_iso


apply_theme()
render_upgrade_css()
current_user = require_login()
operator = current_user["display_name"]

render_page_header("Price Comparison", "Supplier-side cost quotations by Project ID + RFQ Item Ref + Supplier.")
render_upgrade_intro(
    "Supplier Price Comparison",
    "Compare supplier quotations by project and RFQ item. Use this page to see price gaps, quotation completeness, package unit totals and supplier-side quotation risk before client quotation.",
)


# -----------------------------------------------------------------------------
# Display helpers only. Business calculations and import logic remain in services.
# -----------------------------------------------------------------------------


def _txt(value: Any, default: str = "-") -> str:
    if value is None:
        return default
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan", "nat", "null"}:
        return default
    return text


def _blank(value: Any) -> str:
    text = _txt(value, "")
    return "" if text == "-" else text


def _num(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan", "-", "null"}:
        return None
    text = text.replace(",", "").replace("$", "").replace("¥", "").replace("￥", "")
    try:
        return float(text)
    except Exception:
        return None


def _boolish(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "yes", "y", "selected", "recommended"}


def _money(value: Any, currency: Any = None, *, detail: bool = False) -> str:
    number = _num(value)
    if number is None:
        return "-"
    cur = _txt(currency, "").upper()
    prefix = "$" if cur in {"USD", "US$", "$", ""} else f"{cur} "
    decimals = 3 if detail else 2
    return f"{prefix}{number:,.{decimals}f}"


def _pct(value: float | None) -> str:
    return f"{value:.1f}%" if value is not None else "-"


def _safe_key(*parts: Any) -> str:
    raw = "_".join(str(p or "") for p in parts)
    return re.sub(r"[^a-zA-Z0-9_]+", "_", raw)[:180]


def _price(row: dict[str, Any]) -> float | None:
    return _num(row.get("supplier_unit_cost"))


def _quote_date(row: dict[str, Any]) -> str:
    return _txt(row.get("quote_date"), "")


def _group_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (_txt(row.get("project_id")), _txt(row.get("rfq_item_ref")), _blank(row.get("item_option")))


def _item_label(row_or_key: dict[str, Any] | tuple[str, str, str]) -> str:
    if isinstance(row_or_key, tuple):
        _, rfq, opt = row_or_key
        return f"{rfq} / {opt}" if opt else rfq
    rfq = _txt(row_or_key.get("rfq_item_ref"))
    opt = _blank(row_or_key.get("item_option"))
    return f"{rfq} / {opt}" if opt else rfq


def _item_spec_for_group(group: list[dict[str, Any]]) -> str:
    for row in group:
        spec = _txt(row.get("item_spec"), "")
        if spec:
            return spec
    return "-"


def _supplier_label(row: dict[str, Any]) -> str:
    code = _txt(row.get("supplier_code"), "")
    name = _txt(row.get("supplier_name"), "")
    if code and name and code != name:
        return f"{code} — {name}"
    return code or name or "-"


def _is_supplier_matched(row: dict[str, Any]) -> bool:
    return bool(_txt(row.get("supplier_id"), ""))


def _required_quote_fields() -> list[tuple[str, str]]:
    return [
        ("supplier_unit_cost", "Price"),
        ("currency", "Currency"),
        ("moq", "MOQ"),
        ("lead_time", "Lead Time"),
        ("quote_date", "Quote Date"),
        ("supplier_code", "Supplier Code"),
        ("supplier_name", "Supplier Name"),
    ]


def _completeness(row: dict[str, Any]) -> tuple[int, list[str], str]:
    missing = [label for field, label in _required_quote_fields() if not _blank(row.get(field))]
    if not _is_supplier_matched(row):
        missing.append("Supplier Master Match")
    price = _price(row)
    if price is not None and price <= 0:
        missing.append("Price <= 0")
    total = len(_required_quote_fields()) + 2
    score = int(round((total - len(missing)) / total * 100)) if total else 0
    explicit_risk = _txt(row.get("quotation_risk"), "")
    if explicit_risk:
        risk = explicit_risk
    elif price is None or "Supplier Master Match" in missing or "Price <= 0" in missing:
        risk = "High"
    elif len(missing) >= 3:
        risk = "Medium"
    elif missing:
        risk = "Low"
    else:
        risk = "Low"
    return max(0, score), missing, risk


def _needs_review(row: dict[str, Any]) -> bool:
    _, missing, risk = _completeness(row)
    return bool(missing) or str(risk).lower() in {"high", "medium"}


def _review_points(row: dict[str, Any]) -> str:
    _, missing, risk = _completeness(row)
    points = list(missing)
    if str(risk).lower() in {"high", "medium"} and "Risk: " + risk not in points:
        points.append("Risk: " + risk)
    return ", ".join(points) if points else "-"


def _latest_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep the latest quote per Project + RFQ Item Ref + Item Option + Supplier.

    Supplier Quote History still shows all quote dates. Main comparison views use
    latest records so repeated uploads/new quote dates do not duplicate the
    current comparison.
    """
    buckets: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for r in rows:
        supplier = _txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "")
        key = (*_group_key(r), supplier)
        current = buckets.get(key)
        if current is None:
            buckets[key] = r
            continue
        if _quote_date(r) >= _quote_date(current):
            buckets[key] = r
    return list(buckets.values())


def _group_rows(rows: list[dict[str, Any]]) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[_group_key(row)].append(row)
    return dict(sorted(groups.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])))


def _project_options(rows: list[dict[str, Any]]) -> list[str]:
    return sorted({_txt(r.get("project_id"), "") for r in rows if _txt(r.get("project_id"), "")})


def _item_options(rows: list[dict[str, Any]]) -> list[str]:
    return sorted({_txt(r.get("rfq_item_ref"), "") for r in rows if _txt(r.get("rfq_item_ref"), "")})


def _supplier_options(rows: list[dict[str, Any]]) -> list[str]:
    values: set[str] = set()
    for r in rows:
        label = _txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "")
        if label:
            values.add(label)
    return sorted(values)


def _apply_filters(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, str]]:
    st.markdown("### Search and filter")
    c1, c2, c3, c4 = st.columns([1.2, 1.0, 1.0, 1.2])
    search = c1.text_input("Search", placeholder="Project ID, RFQ Item Ref, item spec, supplier, remarks...", key="price_comp_search")
    project = c2.selectbox("Project ID", ["All"] + _project_options(rows), key="price_comp_project")
    item = c3.selectbox("RFQ Item Ref", ["All"] + _item_options(rows), key="price_comp_item")
    supplier = c4.selectbox("Supplier", ["All"] + _supplier_options(rows), key="price_comp_supplier")

    filtered = rows
    if project != "All":
        filtered = [r for r in filtered if _txt(r.get("project_id"), "") == project]
    if item != "All":
        filtered = [r for r in filtered if _txt(r.get("rfq_item_ref"), "") == item]
    if supplier != "All":
        filtered = [r for r in filtered if supplier in {_txt(r.get("supplier_code"), ""), _txt(r.get("supplier_name"), "")}]
    if search:
        keyword = search.strip().lower()
        filtered = [r for r in filtered if any(keyword in str(v or "").lower() for v in r.values())]
    return filtered, {"project": project, "item": item, "supplier": supplier, "search": search or ""}


def _table(data: list[dict[str, Any]], *, key: str, height: int | None = None) -> None:
    if not data:
        st.info("No records to display.")
        return
    df = pd.DataFrame(data)
    for col in df.columns:
        df[col] = df[col].map(lambda v: "-" if v is None else str(v))

    # Streamlit 1.57 does not accept height=None.
    # Only pass height when a real positive integer / supported string is provided.
    kwargs = {"hide_index": True, "width": "stretch", "key": key}
    if isinstance(height, int) and height > 0:
        kwargs["height"] = height
    elif isinstance(height, str) and height in {"auto", "content", "stretch"}:
        kwargs["height"] = height
    st.dataframe(df, **kwargs)


# -----------------------------------------------------------------------------
# Decision persistence. This is intentionally separate from original quote data.
# -----------------------------------------------------------------------------


def _fetch_decisions() -> list[dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    try:
        execute(cur, "SELECT * FROM price_comparison_decisions")
        rows = [dict(r) for r in cur.fetchall()]
    except Exception:
        rows = []
    finally:
        conn.close()
    return rows


def _decision_key(level: str, project_id: str, rfq_item_ref: str = "", item_option: str = "", supplier_quote_id: str = "") -> tuple[str, str, str, str, str]:
    return (level, project_id or "", rfq_item_ref or "", item_option or "", supplier_quote_id or "")


def _decision_map(rows: list[dict[str, Any]]) -> dict[tuple[str, str, str, str, str], dict[str, Any]]:
    return {
        _decision_key(
            _txt(r.get("decision_level"), ""),
            _txt(r.get("project_id"), ""),
            _blank(r.get("rfq_item_ref")),
            _blank(r.get("item_option")),
            _blank(r.get("supplier_quote_id")),
        ): r
        for r in rows
    }


def _save_decision(
    *,
    decision_level: str,
    project_id: str,
    rfq_item_ref: str = "",
    item_option: str = "",
    supplier_quote_id: str = "",
    selected_supplier_code: str = "",
    selection_status: str = "",
    selection_note: str = "",
    review_note: str = "",
) -> None:
    now = now_iso()
    conn = get_connection()
    cur = conn.cursor()
    try:
        execute(
            cur,
            """
            SELECT * FROM price_comparison_decisions
            WHERE decision_level = ? AND project_id = ?
              AND COALESCE(rfq_item_ref, '') = ?
              AND COALESCE(item_option, '') = ?
              AND COALESCE(supplier_quote_id, '') = ?
            LIMIT 1
            """,
            (decision_level, project_id, rfq_item_ref or "", item_option or "", supplier_quote_id or ""),
        )
        existing = cur.fetchone()
        if existing:
            execute(
                cur,
                """
                UPDATE price_comparison_decisions
                SET selected_supplier_code = ?, selection_status = ?, selection_note = ?, review_note = ?, updated_at = ?, updated_by = ?
                WHERE decision_id = ?
                """,
                (selected_supplier_code, selection_status, selection_note, review_note, now, operator, existing["decision_id"]),
            )
        else:
            execute(
                cur,
                """
                INSERT INTO price_comparison_decisions
                (decision_id, decision_level, project_id, rfq_item_ref, item_option, supplier_quote_id,
                 selected_supplier_code, selection_status, selection_note, review_note, created_at, created_by, updated_at, updated_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    f"PCD-{uuid.uuid4().hex[:12].upper()}",
                    decision_level,
                    project_id,
                    rfq_item_ref or None,
                    item_option or None,
                    supplier_quote_id or None,
                    selected_supplier_code or None,
                    selection_status or None,
                    selection_note or None,
                    review_note or None,
                    now,
                    operator,
                    now,
                    operator,
                ),
            )
        conn.commit()
    finally:
        conn.close()


def _selected_item_decision(decisions: dict[tuple[str, str, str, str, str], dict[str, Any]], project_id: str, rfq_item_ref: str, item_option: str) -> dict[str, Any] | None:
    return decisions.get(_decision_key("Item", project_id, rfq_item_ref, item_option, ""))


def _selected_project_decision(decisions: dict[tuple[str, str, str, str, str], dict[str, Any]], project_id: str) -> dict[str, Any] | None:
    return decisions.get(_decision_key("Project", project_id, "", "", ""))


# -----------------------------------------------------------------------------
# Row builders
# -----------------------------------------------------------------------------


def _comparison_table_rows(rows: list[dict[str, Any]], decisions: dict[tuple[str, str, str, str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    sorted_rows = sorted(
        rows,
        key=lambda r: (
            _price(r) is None,
            _price(r) if _price(r) is not None else 10**12,
            _txt(r.get("supplier_code")),
            _txt(r.get("quote_date")),
        ),
    )
    for r in sorted_rows:
        qid = _blank(r.get("supplier_quote_id"))
        review_decision = decisions.get(_decision_key("Quote Review", _txt(r.get("project_id"), ""), _txt(r.get("rfq_item_ref"), ""), _blank(r.get("item_option")), qid))
        output.append(
            {
                "RFQ Item Ref": _txt(r.get("rfq_item_ref")),
                "Item Option": _blank(r.get("item_option")) or "-",
                "Item Spec": _txt(r.get("item_spec")),
                "Supplier Code": _txt(r.get("supplier_code")),
                "Supplier Name": _txt(r.get("supplier_name")),
                "Unit Cost": _money(r.get("supplier_unit_cost"), r.get("currency"), detail=True),
                "Currency": _txt(r.get("currency")),
                "MOQ": _txt(r.get("moq")),
                "Lead Time": _txt(r.get("lead_time")),
                "Price Term": _txt(r.get("price_term")),
                "Tooling Cost": _money(r.get("tooling_cost"), r.get("currency"), detail=True),
                "Quote Date": _txt(r.get("quote_date")),
                "Recommended": "Yes" if _boolish(r.get("recommended_supplier")) else "No",
                "Selected Flag": "Yes" if _boolish(r.get("selected_supplier")) else "No",
                "Need Review": "Yes" if _needs_review(r) else "No",
                "Review Points": _review_points(r),
                "Quote Review Note": _txt(review_decision.get("review_note") if review_decision else None),
                "Remarks": _txt(r.get("remarks")),
            }
        )
    return output


def _build_completeness_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        score, missing, risk = _completeness(r)
        out.append(
            {
                "Project ID": _txt(r.get("project_id")),
                "RFQ Item Ref": _txt(r.get("rfq_item_ref")),
                "Item Option": _blank(r.get("item_option")) or "-",
                "Item Spec": _txt(r.get("item_spec")),
                "Supplier Code": _txt(r.get("supplier_code")),
                "Supplier Name": _txt(r.get("supplier_name")),
                "Unit Cost": _money(r.get("supplier_unit_cost"), r.get("currency"), detail=True),
                "MOQ": _txt(r.get("moq")),
                "Lead Time": _txt(r.get("lead_time")),
                "Quote Date": _txt(r.get("quote_date")),
                "Supplier Matched": "Yes" if _is_supplier_matched(r) else "Review",
                "Completeness": f"{score}%",
                "Risk": risk,
                "Need Review": "Yes" if _needs_review(r) else "No",
                "Missing / Review Points": ", ".join(missing) if missing else "-",
            }
        )
    return out


def _build_saving_rows(rows: list[dict[str, Any]], decisions: dict[tuple[str, str, str, str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for (project_id, rfq_item_ref, item_option), group in _group_rows(_latest_rows(rows)).items():
        priced = [r for r in group if _price(r) is not None]
        item_spec = _item_spec_for_group(group)
        decision = _selected_item_decision(decisions, project_id, rfq_item_ref, item_option)
        if not priced:
            out.append(
                {
                    "Project ID": project_id,
                    "RFQ Item Ref": rfq_item_ref,
                    "Item Option": item_option or "-",
                    "Item Spec": item_spec,
                    "Suppliers": len(group),
                    "Lowest Supplier": "-",
                    "Lowest Price": "-",
                    "Highest Supplier": "-",
                    "Highest Price": "-",
                    "Average Price": "-",
                    "Price Gap": "-",
                    "Saving %": "-",
                    "Need Review": "No valid price",
                    "Comparison Status": "No Quote",
                }
            )
            continue
        prices = [_price(r) for r in priced if _price(r) is not None]
        assert prices
        low = min(prices)
        high = max(prices)
        avg = sum(prices) / len(prices)
        low_row = next(r for r in priced if _price(r) == low)
        high_row = next(r for r in priced if _price(r) == high)
        gap = high - low
        saving = (gap / high * 100) if high else 0
        status = "Selected" if decision and _txt(decision.get("selection_status"), "") == "Selected" else ("Single Quote" if len(priced) == 1 else "Comparable")
        out.append(
            {
                "Project ID": project_id,
                "RFQ Item Ref": rfq_item_ref,
                "Item Option": item_option or "-",
                "Item Spec": item_spec,
                "Suppliers": len({_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "") for r in group}),
                "Lowest Supplier": _supplier_label(low_row),
                "Lowest Price": _money(low, low_row.get("currency")),
                "Highest Supplier": _supplier_label(high_row),
                "Highest Price": _money(high, high_row.get("currency")),
                "Average Price": _money(avg, low_row.get("currency")),
                "Price Gap": _money(gap, low_row.get("currency")),
                "Saving %": _pct(saving) if len(prices) > 1 else "-",
                "Need Review": sum(1 for r in group if _needs_review(r)),
                "Comparison Status": status,
            }
        )
    return out


def _history_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    grouped: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        supplier = _txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "")
        grouped[(*_group_key(r), supplier)].append(r)
    for (project_id, rfq_item_ref, item_option, supplier), group in sorted(grouped.items()):
        sorted_group = sorted(group, key=lambda r: (_quote_date(r), _txt(r.get("supplier_quote_id"))))
        prev_price: float | None = None
        for r in sorted_group:
            price = _price(r)
            change = price - prev_price if price is not None and prev_price is not None else None
            change_pct = (change / prev_price * 100) if change is not None and prev_price else None
            out.append(
                {
                    "Project ID": project_id,
                    "RFQ Item Ref": rfq_item_ref,
                    "Item Option": item_option or "-",
                    "Item Spec": _txt(r.get("item_spec")),
                    "Supplier Code": _txt(r.get("supplier_code")),
                    "Supplier Name": _txt(r.get("supplier_name")),
                    "Quote Date": _txt(r.get("quote_date")),
                    "Unit Cost": _money(price, r.get("currency"), detail=True),
                    "Previous Unit Cost": _money(prev_price, r.get("currency"), detail=True) if prev_price is not None else "-",
                    "Change": _money(change, r.get("currency"), detail=True) if change is not None else "-",
                    "Change %": _pct(change_pct),
                    "MOQ": _txt(r.get("moq")),
                    "Lead Time": _txt(r.get("lead_time")),
                    "Currency": _txt(r.get("currency")),
                    "Remarks": _txt(r.get("remarks")),
                }
            )
            if price is not None:
                prev_price = price
    return out


def _project_summary(rows: list[dict[str, Any]], decisions: dict[tuple[str, str, str, str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    latest = _latest_rows(rows)
    project_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in latest:
        project_groups[_txt(r.get("project_id"))].append(r)
    summaries = []
    for project_id, project_rows in sorted(project_groups.items()):
        item_groups = _group_rows(project_rows)
        item_count = len(item_groups)
        supplier_count = len({_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "") for r in project_rows})
        comparable = 0
        single = 0
        no_quote = 0
        selected_items = 0
        review_item_count = 0
        for (p, rfq, opt), group in item_groups.items():
            priced_suppliers = {_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "") for r in group if _price(r) is not None}
            if not priced_suppliers:
                no_quote += 1
            elif len(priced_suppliers) == 1:
                single += 1
            else:
                comparable += 1
            if _selected_item_decision(decisions, p, rfq, opt):
                selected_items += 1
            if any(_needs_review(r) for r in group):
                review_item_count += 1
        project_decision = _selected_project_decision(decisions, project_id)
        if item_count == 0:
            status = "No Quote"
        elif review_item_count:
            status = "Need Review"
        elif selected_items == item_count or (project_decision and _txt(project_decision.get("selection_status"), "") == "Selected"):
            status = "Selected"
        elif selected_items:
            status = "Partially Selected"
        elif comparable or single:
            status = "In Comparison"
        else:
            status = "No Quote"
        summaries.append(
            {
                "Project ID": project_id,
                "Items": item_count,
                "Suppliers": supplier_count,
                "Comparable Items": comparable,
                "Single Quote Items": single,
                "No Quote Items": no_quote,
                "Selected Items": selected_items,
                "Need Review": review_item_count,
                "Comparison Status": status,
            }
        )
    return summaries


def _project_unit_total_rows(project_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest = _latest_rows(project_rows)
    item_groups = _group_rows(latest)
    all_keys = list(item_groups.keys())
    options = sorted({k[2] for k in all_keys if k[2]})
    packages = options or [""]
    suppliers = sorted({_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "") for r in latest if (_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), ""))})
    out = []
    for package in packages:
        required_keys = [k for k in all_keys if not k[2] or k[2] == package]
        required_count = len(required_keys)
        for supplier in suppliers:
            total = 0.0
            covered = 0
            missing: list[str] = []
            for key in required_keys:
                group = [r for r in item_groups.get(key, []) if (_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "")) == supplier and _price(r) is not None]
                if group:
                    # latest_rows already collapsed supplier/group, so first row is enough.
                    total += float(_price(group[0]) or 0)
                    covered += 1
                else:
                    missing.append(_item_label(key))
            if covered == 0:
                continue
            out.append(
                {
                    "Supplier": supplier,
                    "Item Option Package": package or "Default / No Option",
                    "Items Covered": f"{covered} / {required_count}",
                    "Project Unit Total": _money(total, "USD"),
                    "Status": "Complete" if covered == required_count else "Incomplete",
                    "Missing RFQ Item Ref": ", ".join(missing) if missing else "-",
                }
            )
    return out


def _excel_safe_value(value: Any) -> Any:
    """Convert values to Excel-safe primitives before export.

    Streamlit/Supabase rows may contain timezone-aware datetimes, lists, dicts,
    pandas NA values or other objects. Writing through openpyxl after normalising
    values prevents export from breaking the Price Comparison page.
    """
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (dict, list, tuple, set)):
        return str(value)
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return value


def _write_sheet(ws: Any, rows: list[dict[str, Any]], fallback_columns: list[str] | None = None) -> None:
    columns: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in columns:
                columns.append(key)
    if not columns:
        columns = fallback_columns or ["Note"]
    ws.append(columns)
    if not rows:
        ws.append(["No records"] + [""] * (len(columns) - 1))
        return
    for row in rows:
        ws.append([_excel_safe_value(row.get(col)) for col in columns])


def _export_workbook(filtered_rows: list[dict[str, Any]], all_rows: list[dict[str, Any]], decisions: list[dict[str, Any]]) -> bytes:
    """Build the Price Comparison Excel export without changing database state.

    Use openpyxl directly instead of pandas.ExcelWriter so the workbook always
    has at least one visible sheet, even when the current filter returns no rows.
    """
    from openpyxl import Workbook

    decision_lookup = _decision_map(decisions)
    summary = _project_summary(filtered_rows, decision_lookup)
    comparison_rows: list[dict[str, Any]] = []
    for group in _group_rows(_latest_rows(filtered_rows)).values():
        comparison_rows.extend(_comparison_table_rows(group, decision_lookup))
    risk_rows = _build_completeness_rows(_latest_rows(filtered_rows))
    saving_rows = _build_saving_rows(filtered_rows, decision_lookup)
    history = _history_rows(filtered_rows)

    wb = Workbook()
    readme = wb.active
    readme.title = "Read Me"
    _write_sheet(readme, [
        {"Note": "Project Unit Total is the sum of RFQ Item Ref unit prices. It is not order total."},
        {"Note": "RFQ Item Ref is the Price Comparison grouping reference. Item Option is optional; blank means no option."},
    ])

    sheets: list[tuple[str, list[dict[str, Any]], list[str]]] = [
        ("Project Summary", summary, ["Project ID", "Items", "Suppliers", "Comparable Items", "Single Quote Items", "Need Review", "Comparison Status"]),
        ("Item Supplier Comparison", comparison_rows, ["RFQ Item Ref", "Item Option", "Item Spec", "Supplier Code", "Supplier Name", "Unit Cost", "Currency", "MOQ", "Lead Time", "Quote Date"]),
        ("Completeness Risk", risk_rows, ["Project ID", "RFQ Item Ref", "Item Option", "Item Spec", "Supplier Code", "Supplier Name", "Completeness", "Risk", "Need Review", "Missing / Review Points"]),
        ("Price Difference Saving", saving_rows, ["Project ID", "RFQ Item Ref", "Item Option", "Item Spec", "Suppliers", "Lowest Supplier", "Lowest Price", "Highest Supplier", "Highest Price", "Saving %"]),
        ("Supplier Quote History", history, ["Project ID", "RFQ Item Ref", "Item Option", "Item Spec", "Supplier Code", "Supplier Name", "Quote Date", "Unit Cost", "Previous Unit Cost", "Change", "Change %"]),
        ("Selection Decisions", decisions, ["decision_level", "project_id", "rfq_item_ref", "item_option", "supplier_quote_id", "selected_supplier_code", "selection_status", "selection_note", "review_note"]),
        ("Raw Records", filtered_rows, ["project_id", "rfq_item_ref", "item_option", "item_spec", "supplier_code", "supplier_name", "supplier_unit_cost", "currency", "quote_date"]),
    ]
    for title, sheet_rows, fallback_columns in sheets:
        ws = wb.create_sheet(title=title[:31])
        _write_sheet(ws, sheet_rows, fallback_columns)

    # Freeze headers and enable filters for easier Excel review.
    for ws in wb.worksheets:
        ws.sheet_state = "visible"
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# -----------------------------------------------------------------------------
# Data
# -----------------------------------------------------------------------------


rows = list_module_records("Supplier Price Comparison", limit=10000)
decision_rows = _fetch_decisions()
decisions = _decision_map(decision_rows)
latest_rows = _latest_rows(rows)
summary_rows = _project_summary(rows, decisions)

selected = sum(1 for d in decision_rows if _txt(d.get("selection_status"), "") == "Selected")
recommended = sum(1 for r in rows if _boolish(r.get("recommended_supplier")))
projects = len({r.get("project_id") for r in rows if r.get("project_id")})
items = len({(r.get("project_id"), r.get("rfq_item_ref"), _blank(r.get("item_option"))) for r in rows if r.get("project_id") and r.get("rfq_item_ref")})
suppliers = len({_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "") for r in rows if (_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), ""))})
render_metric_grid(
    {
        "Quote Records": len(rows),
        "Projects": projects,
        "RFQ Groups": items,
        "Suppliers": suppliers,
        "Selected Decisions": selected,
        "Recommended Flags": recommended,
        "Need Review Items": sum(1 for r in summary_rows if str(r.get("Need Review")) not in {"0", "-"}),
        "Comparable Items": sum(int(r.get("Comparable Items") or 0) for r in summary_rows),
    }
)

# Export buttons are read-only.
filtered, filter_state = _apply_filters(rows)
export_cols = st.columns([1, 1, 3])
selected_project_rows = filtered if filter_state.get("project") != "All" else []
if selected_project_rows:
    export_cols[0].download_button(
        "Export Selected Project",
        data=_export_workbook(selected_project_rows, rows, decision_rows),
        file_name=f"price_comparison_{filter_state.get('project')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
else:
    export_cols[0].button("Export Selected Project", disabled=True, help="Select one Project ID first.", width="stretch")
export_cols[1].download_button(
    "Export All Price Comparison",
    data=_export_workbook(rows, rows, decision_rows),
    file_name=f"price_comparison_all_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    width="stretch",
)

with st.expander("Add supplier quote", expanded=False):
    with st.form("supplier_quote_form"):
        c1, c2, c3, c4 = st.columns(4)
        project_id = c1.text_input("Project ID")
        rfq_item_ref = c2.text_input("RFQ Item Ref")
        item_option = c3.text_input("Item Option", help="Optional. Blank means no option.")
        quote_round = c4.text_input("Quote Round", value="1")
        item_spec = st.text_area("Item Spec", height=80)
        c1, c2, c3 = st.columns(3)
        supplier_code = c1.text_input("Supplier Code")
        supplier_name = c2.text_input("Supplier Name")
        quote_date = c3.date_input("Quote Date", value=None)
        c1, c2, c3, c4 = st.columns(4)
        supplier_unit_cost = c1.number_input("Supplier Unit Cost", min_value=0.0, step=0.001, format="%.3f", value=0.0)
        currency = c2.text_input("Currency", value="USD")
        price_term = c3.text_input("Price Term")
        lead_time = c4.text_input("Lead Time")
        c1, c2 = st.columns(2)
        recommended_supplier = c1.checkbox("Recommended Supplier")
        selected_supplier = c2.checkbox("Selected Supplier")
        with st.expander("Risk / missing info", expanded=False):
            quotation_quality = st.selectbox("Quotation Quality", ["", "Complete", "Partial", "Poor"])
            quotation_risk = st.selectbox("Quotation Risk", ["", "Low", "Medium", "High"])
            missing_info = st.text_area("Missing Information", height=80)
            selection_reason = st.text_area("Selection Reason", height=80)
        remarks = st.text_area("Remarks", height=80)
        submitted = st.form_submit_button("Save Supplier Quote", type="primary")
        if submitted:
            if not project_id.strip() or not rfq_item_ref.strip() or not supplier_name.strip():
                st.error("Project ID, RFQ Item Ref and Supplier Name are required.")
            else:
                upsert_module_record(
                    "Supplier Price Comparison",
                    {
                        "project_id": project_id,
                        "rfq_item_ref": rfq_item_ref,
                        "item_option": item_option,
                        "item_spec": item_spec,
                        "supplier_code": supplier_code,
                        "supplier_name": supplier_name,
                        "quote_round": quote_round,
                        "quote_date": quote_date.isoformat() if quote_date else None,
                        "supplier_unit_cost": supplier_unit_cost,
                        "currency": currency,
                        "price_term": price_term,
                        "lead_time": lead_time,
                        "recommended_supplier": recommended_supplier,
                        "selected_supplier": selected_supplier,
                        "quotation_quality": quotation_quality,
                        "quotation_risk": quotation_risk,
                        "missing_info": missing_info,
                        "selection_reason": selection_reason,
                        "remarks": remarks,
                    },
                    operator=operator,
                )
                st.success("Supplier quote saved.")
                st.rerun()

st.markdown("---")
view_tab, risk_tab, saving_tab, history_tab, raw_tab = st.tabs(
    [
        "1. Project / Item / Supplier Comparison",
        "2. Quote Completeness / Risk",
        "3. Price Difference / Saving",
        "4. Supplier Quote History",
        "Raw Records",
    ]
)

with view_tab:
    st.markdown("### Project → RFQ Item Ref → Supplier comparison")
    st.caption("Project Unit Total is the sum of RFQ Item Ref unit prices. It is not an order total.")
    project_buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in _latest_rows(filtered):
        project_buckets[_txt(r.get("project_id"))].append(r)
    if not project_buckets:
        st.info("No matching supplier quote records.")
    for project_id, project_rows in sorted(project_buckets.items()):
        project_summary = next((s for s in _project_summary(project_rows, decisions) if s["Project ID"] == project_id), None)
        if project_summary:
            header = (
                f"{project_id} — Items: {project_summary['Items']} | Suppliers: {project_summary['Suppliers']} | "
                f"Comparable Items: {project_summary['Comparable Items']} | Single Quote Items: {project_summary['Single Quote Items']} | "
                f"Need Review: {project_summary['Need Review']} | Status: {project_summary['Comparison Status']}"
            )
        else:
            header = project_id
        with st.expander(header, expanded=(filter_state.get("project") == project_id)):
            st.markdown("#### Project Unit Total by Supplier")
            _table(_project_unit_total_rows(project_rows), key=f"project_total_{_safe_key(project_id)}")

            st.markdown("#### Project-level selection")
            project_decision = _selected_project_decision(decisions, project_id) or {}
            suppliers_for_project = ["-"] + sorted({_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "") for r in project_rows})
            with st.form(f"project_decision_{_safe_key(project_id)}"):
                c1, c2 = st.columns([1, 1])
                supplier_choice = c1.selectbox("Selected supplier / package", suppliers_for_project, index=suppliers_for_project.index(_txt(project_decision.get("selected_supplier_code"), "-")) if _txt(project_decision.get("selected_supplier_code"), "-") in suppliers_for_project else 0)
                status_choice = c2.selectbox("Project comparison status", ["In Comparison", "Partially Selected", "Selected", "Need Review"], index=0)
                note = st.text_area("Project selection note", value=_txt(project_decision.get("selection_note"), ""), height=70)
                if st.form_submit_button("Save Project Selection"):
                    _save_decision(decision_level="Project", project_id=project_id, selected_supplier_code="" if supplier_choice == "-" else supplier_choice, selection_status=status_choice, selection_note=note)
                    st.success("Project selection saved.")
                    st.rerun()

            for (p, rfq, opt), group in _group_rows(project_rows).items():
                prices = [_price(r) for r in group if _price(r) is not None]
                supplier_count = len({_txt(r.get("supplier_code"), "") or _txt(r.get("supplier_name"), "") for r in group})
                lowest = min(prices) if prices else None
                highest = max(prices) if prices else None
                gap_pct = ((highest - lowest) / highest * 100) if lowest is not None and highest else None
                need_review_count = sum(1 for r in group if _needs_review(r))
                item_title = f"{_item_label((p, rfq, opt))} — {_item_spec_for_group(group)}"
                item_caption = (
                    f"Suppliers: {supplier_count} | Lowest: {_money(lowest, group[0].get('currency'))} | "
                    f"Highest: {_money(highest, group[0].get('currency'))} | Gap: {_pct(gap_pct) if len(prices) > 1 else '-'} | "
                    f"Need Review: {need_review_count}"
                )
                with st.expander(f"{item_title} — {item_caption}", expanded=False):
                    item_decision = _selected_item_decision(decisions, p, rfq, opt) or {}
                    quote_labels = ["-"] + [f"{_txt(r.get('supplier_code'), '') or _txt(r.get('supplier_name'), '')} | {_money(r.get('supplier_unit_cost'), r.get('currency'), detail=True)} | {_txt(r.get('quote_date'))}" for r in group]
                    quote_map = {label: r for label, r in zip(quote_labels[1:], group)}
                    with st.form(f"item_decision_{_safe_key(p, rfq, opt)}"):
                        c1, c2 = st.columns([1.2, 1.0])
                        selected_label = c1.selectbox("Selected supplier quote", quote_labels)
                        item_status = c2.selectbox("RFQ Item Ref status", ["No Quote", "Single Quote", "Comparable", "Selected", "Need Review"], index=3 if _txt(item_decision.get("selection_status"), "") == "Selected" else 2)
                        item_note = st.text_area("Item selection note", value=_txt(item_decision.get("selection_note"), ""), height=70)
                        if st.form_submit_button("Save Item Selection"):
                            chosen = quote_map.get(selected_label)
                            _save_decision(
                                decision_level="Item",
                                project_id=p,
                                rfq_item_ref=rfq,
                                item_option=opt,
                                selected_supplier_code=(_txt(chosen.get("supplier_code"), "") if chosen else ""),
                                supplier_quote_id="",
                                selection_status=item_status,
                                selection_note=item_note,
                            )
                            st.success("Item selection saved.")
                            st.rerun()

                    with st.expander("Quote review note for unselected / reference quotations", expanded=False):
                        with st.form(f"quote_review_{_safe_key(p, rfq, opt)}"):
                            q_label = st.selectbox("Supplier quote", quote_labels, key=f"qreview_select_{_safe_key(p, rfq, opt)}")
                            review_note = st.text_area("Review note", height=70, placeholder="Example: price too high / MOQ not acceptable / lead time too long / only for reference")
                            if st.form_submit_button("Save Quote Review Note"):
                                chosen = quote_map.get(q_label)
                                if chosen:
                                    _save_decision(
                                        decision_level="Quote Review",
                                        project_id=p,
                                        rfq_item_ref=rfq,
                                        item_option=opt,
                                        supplier_quote_id=_blank(chosen.get("supplier_quote_id")),
                                        selected_supplier_code=_txt(chosen.get("supplier_code"), ""),
                                        selection_status="Reviewed",
                                        review_note=review_note,
                                    )
                                    st.success("Quote review note saved.")
                                    st.rerun()
                                else:
                                    st.warning("Please select a supplier quote first.")
                    _table(_comparison_table_rows(group, decisions), key=f"compare_{_safe_key(p, rfq, opt)}", height=360)

with risk_tab:
    st.markdown("### Quote completeness and risk view")
    st.caption("Need Review counts supplier quote rows with missing/abnormal fields or supplier master mismatch.")
    risk_rows = _build_completeness_rows(_latest_rows(filtered))
    render_metric_grid({"Rows Checked": len(risk_rows), "Need Review": sum(1 for r in risk_rows if r.get("Need Review") == "Yes"), "High Risk": sum(1 for r in risk_rows if r.get("Risk") == "High"), "Ready / Low Risk": sum(1 for r in risk_rows if r.get("Risk") == "Low")})
    _table(risk_rows, key="price_risk_table", height=520)

with saving_tab:
    st.markdown("### Price difference and saving view")
    st.caption("Highlights lowest supplier, highest supplier and potential saving range for each Project ID + RFQ Item Ref + Item Option.")
    saving_rows = _build_saving_rows(filtered, decisions)
    _table(saving_rows, key="price_saving_table", height=520)

with history_tab:
    st.markdown("### Supplier Quote History")
    st.caption("Shows the same supplier's quote changes over time for the same Project ID + RFQ Item Ref + Item Option.")
    _table(_history_rows(filtered), key="supplier_quote_history_table", height=560)

with raw_tab:
    st.markdown("### Raw supplier price records")
    st.caption("Original extension records are kept here for audit and troubleshooting.")
    render_layered_records(
        "Supplier Price Comparison",
        filtered,
        key_prefix="price_page",
        summary_field="comparison_status",
        preview_columns=[
            "project_id",
            "rfq_item_ref",
            "item_option",
            "item_spec",
            "supplier_code",
            "supplier_name",
            "quote_date",
            "supplier_unit_cost",
            "currency",
            "moq",
            "lead_time",
            "recommended_supplier",
            "selected_supplier",
            "comparison_status",
        ],
    )
