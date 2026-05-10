from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROCESS_DEFINITIONS: dict[str, dict[str, Any]] = {
    "QP-01": {
        "process_code": "QP-01",
        "process_name": "Request for Quotation (RFQ) Requirement Control",
        "short_name": "RFQ Requirement Control",
        "process_type": "Formal Quality Process",
        "version": "V1.0",
        "status": "Active",
        "owner": "Harley",
        "quality_owner": "Harley",
        "business_owner": "Maria",
        "final_approver": "Ehab",
        "effective_date": "Pending approval",
        "scope": "Normal bulk order process, especially key customers such as EHS / Keter.",
        "purpose": "Upgrade the existing RFQ Working File into an RFQ Control Layer: keep free notes and file links, while adding RFQ status, missing information, risk level, owner, next step, due date and requirement checklist.",
        "trigger": "A customer RFQ or confirmed quotation requirement is received and a project working file is opened.",
        "existing_sources": "Existing RFQ Working File, Sales Board, Project Details, Supplier Details, Price Comparison, Meeting Mode",
        "extension_needed": "rfq_requirement_control extension table + RFQ Working File mapping",
    },
    "QP-02": {
        "process_code": "QP-02",
        "process_name": "Sample Control",
        "short_name": "Sample Control",
        "process_type": "Formal Quality Process",
        "version": "V1.0",
        "status": "Draft",
        "owner": "Harley",
        "quality_owner": "Harley",
        "business_owner": "Maria",
        "final_approver": "Ehab if high risk",
        "effective_date": "Pending approval",
        "scope": "Customer sample, testing sample, retained sample and golden/reference sample for normal bulk orders.",
        "purpose": "Control sample purpose, sample quantity, sample checking, client feedback, re-sample decision, sample retention and disposal.",
        "trigger": "A new normal order, new product, first delivery or customer sample request is identified.",
        "existing_sources": "Sample Tracking, Sales / Project Details, Operation / Order Details, Supplier Details, Meeting Mode",
        "extension_needed": "sample_control_extension",
    },
    "QP-03": {
        "process_code": "QP-03",
        "process_name": "Order Setup Control",
        "short_name": "Order Setup Control",
        "process_type": "Formal Quality Process",
        "version": "V1.0",
        "status": "Draft",
        "owner": "Harley",
        "quality_owner": "Harley",
        "business_owner": "Maria",
        "final_approver": "Ehab if high risk",
        "effective_date": "Pending approval",
        "scope": "Customer-confirmed normal orders before Purchase Order (PO) release to supplier.",
        "purpose": "Control order folder, Order Excel input, quantity check, supplier price check, client price/margin check, quality requirements and PO release status.",
        "trigger": "Client confirms order and internal order folder / Order Excel needs to be created.",
        "existing_sources": "Operation Board, Order Details, Project Details, WeCom folder link, Meeting Mode",
        "extension_needed": "order_setup_control",
    },
    "QP-04": {
        "process_code": "QP-04",
        "process_name": "Inspection & Shipment Release Control",
        "short_name": "Inspection & Shipment Release",
        "process_type": "Formal Quality Process",
        "version": "V1.0",
        "status": "Draft",
        "owner": "Harley",
        "quality_owner": "Harley",
        "business_owner": "Maria",
        "final_approver": "Ehab if fail / unresolved / high risk",
        "effective_date": "Pending approval",
        "scope": "Supplier self-inspection, COA, COC, Mark inspection, third-party inspection, customer approval and shipment release.",
        "purpose": "Control inspection trigger rules, required documents, inspection results, deviation handling, customer approval and shipment release decision.",
        "trigger": "Supplier reports goods ready, shipment is planned, inspection is required, or a customer quality rule applies.",
        "existing_sources": "Order Details, Operation Board, Supplier Details, Project Details, Meeting Mode",
        "extension_needed": "inspection_shipment_release_control",
    },
    "QP-05": {
        "process_code": "QP-05",
        "process_name": "Quality Complaint Closure",
        "short_name": "Quality Complaint Closure",
        "process_type": "Formal Quality Process",
        "version": "V1.0",
        "status": "Draft",
        "owner": "Harley",
        "quality_owner": "Harley",
        "business_owner": "Maria for commercial settlement",
        "final_approver": "Ehab if serious / unresolved / compensation",
        "effective_date": "Pending approval",
        "scope": "Product quality issue, inspection failure, packaging quality issue and customer quality complaint after shipment.",
        "purpose": "Control complaint record, evidence, root cause, corrective action, preventive action, customer response, closure evidence and supplier history update.",
        "trigger": "A quality issue, inspection failure or customer complaint is received.",
        "existing_sources": "Project Details, Operation / Order Details, Supplier Details, Meeting Mode, Supplier Management",
        "extension_needed": "quality_complaint_capa",
    },
    "SV-01": {
        "process_code": "SV-01",
        "process_name": "Supplier Management",
        "short_name": "Supplier Management",
        "process_type": "Supporting View / Management Module",
        "version": "V1.0",
        "status": "Draft",
        "owner": "Harley",
        "quality_owner": "Harley",
        "business_owner": "Maria",
        "final_approver": "Ehab for hold / strategic decision",
        "effective_date": "Pending approval",
        "scope": "Supplier master data, quotation history, order history, sample performance, inspection performance, complaints and quality risk summary.",
        "purpose": "Summarise supplier performance and risk without changing Supplier Details core business logic.",
        "trigger": "Supplier is used for RFQ, sample, order, inspection or quality complaint.",
        "existing_sources": "Supplier Details, Price Comparison, Order Details, Sample Tracking, Inspection Control, Complaint / CAPA",
        "extension_needed": "supplier_quality_summary / derived view",
    },
}

PROCESS_ORDER = ["QP-01", "QP-02", "QP-03", "QP-04", "QP-05", "SV-01"]

CONTROL_POINTS: dict[str, list[dict[str, str]]] = {
    "QP-01": [
        {"step": "1", "control_point": "Receive RFQ and open RFQ Working File", "owner": "Ehab & Maria / Operator", "required_record": "RFQ ID / RFQ Working File link / customer / product", "risk_controlled": "RFQ not captured or project file entry point is unclear"},
        {"step": "2", "control_point": "Record original requirement notes and project file links", "owner": "Operator / Sandy", "required_record": "Original notes + sourcing / sampling / design / quotation links", "risk_controlled": "Project information and documents are scattered"},
        {"step": "3", "control_point": "Complete RFQ Control Summary and requirement checklist", "owner": "Harley + Maria + Sandy", "required_record": "RFQ status / missing information / checklist / owner / next step / due date", "risk_controlled": "Working file is readable but not controllable"},
        {"step": "4", "control_point": "Collect supplier quotation", "owner": "Sandy", "required_record": "Supplier quote status", "risk_controlled": "Supplier cost / lead time not clear"},
        {"step": "5", "control_point": "Quality & compliance risk review", "owner": "Harley", "required_record": "Quality/compliance risk level", "risk_controlled": "Compliance, testing, inspection or quality requirement missed"},
        {"step": "6", "control_point": "Commercial & business risk review", "owner": "Maria", "required_record": "Commercial risk level", "risk_controlled": "Hidden cost, delivery, margin or customer responsibility risk"},
        {"step": "7", "control_point": "Final gate decision", "owner": "Ehab", "required_record": "Final risk level and action", "risk_controlled": "High-risk RFQ continues without management decision"},
    ],
    "QP-02": [
        {"step": "1", "control_point": "Decide whether sample is required", "owner": "Harley + Sandy", "required_record": "Sample required?", "risk_controlled": "New product proceeds without sample control"},
        {"step": "2", "control_point": "Confirm customer sample quantity", "owner": "Maria + Harley", "required_record": "Customer sample qty", "risk_controlled": "Insufficient customer samples"},
        {"step": "3", "control_point": "Confirm testing sample quantity", "owner": "Harley + Mark", "required_record": "Testing sample qty", "risk_controlled": "Testing cannot be completed due to insufficient samples"},
        {"step": "4", "control_point": "Confirm retained sample quantity", "owner": "Harley + Mark", "required_record": "Retained sample qty / location", "risk_controlled": "No reference sample for inspection or complaint"},
        {"step": "5", "control_point": "Integrate sample plan and contact supplier", "owner": "Sandy", "required_record": "Final sample plan", "risk_controlled": "Multiple people give supplier inconsistent sample instructions"},
        {"step": "6", "control_point": "Check received sample", "owner": "Mark", "required_record": "Sample check result", "risk_controlled": "Failed sample sent to customer"},
        {"step": "7", "control_point": "Record customer feedback and close / re-sample", "owner": "Maria + Harley + Mark", "required_record": "Client feedback / re-sample decision", "risk_controlled": "Sample approval history missing"},
    ],
    "QP-03": [
        {"step": "1", "control_point": "Create WeCom order folder", "owner": "Candy", "required_record": "Order folder link", "risk_controlled": "Order documents scattered"},
        {"step": "2", "control_point": "Upload blank Order Excel", "owner": "Candy", "required_record": "Blank Order Excel uploaded", "risk_controlled": "No controlled order data entry file"},
        {"step": "3", "control_point": "Input order information", "owner": "Sophia", "required_record": "Input completed status", "risk_controlled": "Order information incomplete"},
        {"step": "4", "control_point": "Check quantity and supplier price", "owner": "Sandy / Camille", "required_record": "Quantity / supplier price check status", "risk_controlled": "Wrong quantity or supplier cost sent to supplier"},
        {"step": "5", "control_point": "Check client price / margin", "owner": "Maria", "required_record": "Client price check status", "risk_controlled": "Margin or customer price error"},
        {"step": "6", "control_point": "Check quality / sample / inspection requirements", "owner": "Harley", "required_record": "Quality requirement check status", "risk_controlled": "EHS quality requirement missed in order setup"},
        {"step": "7", "control_point": "Release PO to supplier", "owner": "Sandy / Camille", "required_record": "PO sent date / supplier confirmation", "risk_controlled": "PO released without required checks"},
    ],
    "QP-04": [
        {"step": "1", "control_point": "Confirm inspection trigger", "owner": "Harley + Mark", "required_record": "Inspection required / trigger reason", "risk_controlled": "Required inspection missed"},
        {"step": "2", "control_point": "Collect supplier self-inspection and COA / COC", "owner": "Sandy", "required_record": "Supplier inspection / COA / COC status", "risk_controlled": "Shipment without required supplier documents"},
        {"step": "3", "control_point": "Arrange Mark or third-party inspection", "owner": "Mark / Maria", "required_record": "Inspection date / inspector", "risk_controlled": "Inspection not scheduled in time"},
        {"step": "4", "control_point": "Review inspection report", "owner": "Mark + Harley", "required_record": "Inspection result / report link", "risk_controlled": "Report not reviewed before shipment"},
        {"step": "5", "control_point": "Handle fail / deviation", "owner": "Harley → Maria / Ehab", "required_record": "Deviation / corrective action", "risk_controlled": "Failed goods released without decision"},
        {"step": "6", "control_point": "Customer approval and shipment release", "owner": "Maria + Candy", "required_record": "Customer approval / release status", "risk_controlled": "Shipment released before customer confirmation"},
    ],
    "QP-05": [
        {"step": "1", "control_point": "Record complaint", "owner": "Harley", "required_record": "Complaint ID / issue description", "risk_controlled": "Complaint handled without formal record"},
        {"step": "2", "control_point": "Collect evidence", "owner": "Harley + Mark", "required_record": "Evidence link", "risk_controlled": "No evidence for supplier/customer discussion"},
        {"step": "3", "control_point": "Analyse root cause", "owner": "Harley + Supplier", "required_record": "Root cause", "risk_controlled": "Repeated problem without cause analysis"},
        {"step": "4", "control_point": "Define Corrective Action and Preventive Action (CAPA)", "owner": "Harley", "required_record": "Corrective and preventive action", "risk_controlled": "No long-term improvement"},
        {"step": "5", "control_point": "Handle commercial settlement if needed", "owner": "Maria / Ehab", "required_record": "Commercial decision", "risk_controlled": "Compensation or replacement responsibility unclear"},
        {"step": "6", "control_point": "Close complaint and update supplier history", "owner": "Harley", "required_record": "Closure evidence / supplier history updated", "risk_controlled": "Complaint closed without evidence or supplier impact"},
    ],
    "SV-01": [
        {"step": "1", "control_point": "Maintain supplier master data", "owner": "Harley", "required_record": "Supplier Details", "risk_controlled": "Supplier information incomplete"},
        {"step": "2", "control_point": "Summarise quotation and order activity", "owner": "System derived / Harley review", "required_record": "Quote count / order count", "risk_controlled": "Supplier activity not visible"},
        {"step": "3", "control_point": "Summarise sample and inspection performance", "owner": "System derived / Harley review", "required_record": "Sample / inspection pass rate", "risk_controlled": "Supplier quality history not used"},
        {"step": "4", "control_point": "Track complaint and open issue", "owner": "Harley", "required_record": "Complaint count / open issue", "risk_controlled": "Repeated supplier issue not escalated"},
        {"step": "5", "control_point": "Define supplier risk and next action", "owner": "Harley + Maria / Ehab if high risk", "required_record": "Supplier risk level / next action", "risk_controlled": "Risk supplier continues without management control"},
    ],
}

QUALITY_TEMPLATE_FIELDS: dict[str, list[dict[str, str]]] = {
    "QP-01": [
        ("rfq_id", "RFQ ID", "Optional", "Text", "No", "Extension", "Unique RFQ requirement control record. Leave blank to let the system generate one."),
        ("project_id", "Project ID", "Recommended", "Text", "Controlled", "Sales Project", "Linked project ID."),
        ("customer", "Customer", "Yes", "Text", "Controlled", "Sales Project", "Customer or client code."),
        ("product_description", "Product Description", "Yes", "Text", "Yes", "Sales / Extension", "Product description or RFQ item summary."),
        ("rfq_working_file_link", "RFQ Working File Link", "Recommended", "Link", "Yes", "Existing RFQ Working File", "Main project working file link used by the team."),
        ("customer_original_request_link", "Customer Original Request Link", "Recommended", "Link", "Yes", "Existing RFQ Working File", "Customer email / original RFQ / source file link."),
        ("sourcing_link", "Sourcing Link", "Recommended", "Link", "Yes", "Existing RFQ Working File", "Supplier sourcing / quotation folder link."),
        ("sampling_link", "Sampling Link", "No", "Link", "Yes", "Existing RFQ Working File", "Sampling folder or sample record link."),
        ("design_file_link", "Design File Link", "Recommended", "Link", "Yes", "Existing RFQ Working File", "Drawing / design file link."),
        ("quotation_to_client_link", "Quotation to Client Link", "No", "Link", "Yes", "Existing RFQ Working File", "Client quotation file link."),
        ("original_requirement_notes", "Original Requirement Notes", "Recommended", "Long Text", "Yes", "Existing RFQ Working File", "Free-text customer requirement notes copied from the working file."),
        ("rfq_received_date", "RFQ Received Date", "Recommended", "Date", "Yes", "Extension", "Date RFQ was received."),
        ("rfq_received_by", "RFQ Received By", "Recommended", "Text", "Yes", "Extension", "Person who received RFQ."),
        ("drawing_received", "Drawing Received", "Recommended", "Yes/No", "Yes", "Extension", "Whether drawing was received."),
        ("specification_received", "Specification Received", "Recommended", "Yes/No", "Yes", "Extension", "Whether specification was received."),
        ("quantity_confirmed", "Quantity Confirmed", "Recommended", "Yes/No", "Yes", "Extension", "Whether quantity is confirmed."),
        ("target_price_received", "Target Price Received", "No", "Yes/No", "Yes", "Extension", "Whether target price exists."),
        ("delivery_requirement", "Delivery Requirement", "No", "Text / Date", "Yes", "Extension", "Customer delivery requirement."),
        ("packaging_requirement", "Packaging Requirement", "No", "Text", "Yes", "Extension", "Packaging requirement."),
        ("testing_requirement", "Testing Requirement", "No", "Text", "Yes", "Extension", "Testing requirement."),
        ("compliance_requirement", "Compliance Requirement", "No", "Text", "Yes", "Extension", "Compliance requirement."),
        ("sample_required", "Sample Required", "Recommended", "Yes/No", "Yes", "Extension", "Whether sample is required."),
        ("inspection_required", "Inspection Required", "Recommended", "Yes/No/By Case", "Yes", "Extension", "Whether inspection is required."),
        ("missing_information", "Missing Information", "No", "Text", "Yes", "Extension", "Missing RFQ information."),
        ("quality_compliance_risk", "Quality / Compliance Risk", "No", "Text", "Yes", "Extension", "Quality and compliance risk identified by Harley."),
        ("commercial_business_risk", "Commercial / Business Risk", "No", "Text", "Yes", "Extension", "Commercial risk identified by Maria."),
        ("harley_review_status", "Harley Review Status", "Recommended", "Status", "Yes", "Extension", "Quality/compliance review status."),
        ("maria_review_status", "Maria Review Status", "Recommended", "Status", "Yes", "Extension", "Business review status."),
        ("current_owner", "Current Owner", "Recommended", "Text", "Yes", "Extension", "Current follow-up owner for this RFQ."),
        ("next_step", "Next Step", "Recommended", "Text", "Yes", "Extension", "Next action for RFQ progress."),
        ("due_date", "Due Date", "Recommended", "Date", "Yes", "Extension", "Due date for the next action."),
        ("risk_level", "Risk Level", "Recommended", "Status", "Yes", "Extension", "Low / Medium / High / Critical."),
        ("ehab_final_decision", "Ehab Final Decision", "No", "Text", "Yes", "Extension", "Final risk action decision."),
        ("rfq_gate_status", "RFQ Gate Status", "Recommended", "Status", "Yes", "Extension", "Open / Pending Information / Ready / Hold / Closed."),
    ],
    "QP-02": [
        ("sample_id", "Sample ID", "Yes", "Text", "No", "Extension", "Unique sample control record."),
        ("project_id", "Project ID", "Yes", "Text", "Controlled", "Sales / Project", "Linked project ID."),
        ("order_no", "Order No", "By case", "Text", "Controlled", "Operation / Order Details", "Linked order number."),
        ("customer", "Customer", "Yes", "Text", "Controlled", "Existing system", "Customer code/name."),
        ("supplier_code", "Supplier Code", "Recommended", "Text", "Controlled", "Supplier Details", "Supplier code."),
        ("supplier_name", "Supplier Name", "Recommended", "Text", "No", "Supplier Details", "Supplier name."),
        ("sample_required", "Sample Required", "Yes", "Yes/No", "Yes", "Extension", "Whether sample is required."),
        ("sample_type", "Sample Type", "Yes", "Text", "Yes", "Extension", "Customer / Testing / Retained / Golden Sample."),
        ("customer_sample_qty", "Customer Sample Qty", "No", "Number", "Yes", "Extension", "Quantity for customer approval."),
        ("testing_sample_qty", "Testing Sample Qty", "No", "Number", "Yes", "Extension", "Quantity for testing."),
        ("retained_sample_qty", "Retained Sample Qty", "No", "Number", "Yes", "Extension", "Quantity retained internally."),
        ("sample_qty_suggested_by", "Sample Qty Suggested By", "No", "Text", "Yes", "Extension", "Person who suggested sample quantity."),
        ("sample_qty_confirmed_by", "Sample Qty Confirmed By", "No", "Text", "Yes", "Extension", "Person who confirmed sample quantity."),
        ("sample_plan_owner", "Sample Plan Owner", "Yes", "Text", "Yes", "Extension", "Person who consolidates sample plan, normally Sandy."),
        ("supplier_sent_date", "Supplier Sent Date", "No", "Date", "Yes", "Extension", "Supplier sample sending date."),
        ("sample_received_date", "Sample Received Date", "No", "Date", "Yes", "Extension", "Sample received date."),
        ("checked_by", "Checked By", "No", "Text", "Yes", "Extension", "Person who checked sample, normally Mark."),
        ("sample_check_result", "Sample Check Result", "No", "Status", "Yes", "Extension", "Pass / Fail / Conditional."),
        ("sent_to_client_date", "Sent to Client Date", "No", "Date", "Yes", "Extension", "Date sample sent to client."),
        ("client_feedback", "Client Feedback", "No", "Text", "Yes", "Extension", "Client sample feedback."),
        ("resample_required", "Re-sample Required", "No", "Yes/No", "Yes", "Extension", "Whether re-sample is required."),
        ("sample_location", "Sample Location", "No", "Text", "Yes", "Extension", "Retained sample location."),
        ("retention_until", "Retention Until", "No", "Date", "Yes", "Extension", "Retention due date."),
        ("disposal_status", "Disposal Status", "No", "Status", "Yes", "Extension", "Retained / Disposed / Pending."),
        ("sample_status", "Sample Status", "Yes", "Status", "Yes", "Extension", "Open / Waiting Supplier / Received / Checked / Sent / Approved / Rejected / Closed."),
    ],
    "QP-03": [
        ("order_setup_id", "Order Setup ID", "Yes", "Text", "No", "Extension", "Unique order setup control record."),
        ("project_id", "Project ID", "Yes", "Text", "Controlled", "Existing system", "Linked project ID."),
        ("order_no", "Order No", "Yes", "Text", "Controlled", "Operation / Order Details", "Order number."),
        ("customer", "Customer", "Yes", "Text", "Controlled", "Existing system", "Customer code/name."),
        ("order_folder_link", "Order Folder Link", "Recommended", "Link", "Yes", "Existing + Extension", "WeCom order folder link."),
        ("blank_order_excel_uploaded", "Blank Order Excel Uploaded", "Yes", "Yes/No", "Yes", "Extension", "Whether blank Order Excel was uploaded."),
        ("order_input_by", "Order Input By", "Yes", "Text", "Yes", "Extension", "Person who input order data, normally Sophia."),
        ("order_input_completed", "Order Input Completed", "Yes", "Yes/No", "Yes", "Extension", "Whether order input is completed."),
        ("quantity_checked_by", "Quantity Checked By", "No", "Text", "Yes", "Extension", "Quantity checker."),
        ("quantity_check_status", "Quantity Check Status", "Yes", "Status", "Yes", "Extension", "Pending / Checked / Issue."),
        ("supplier_price_checked_by", "Supplier Price Checked By", "No", "Text", "Yes", "Extension", "Supplier price checker."),
        ("supplier_price_check_status", "Supplier Price Check Status", "Yes", "Status", "Yes", "Extension", "Pending / Checked / Issue."),
        ("client_price_checked_by", "Client Price Checked By", "No", "Text", "Yes", "Extension", "Client price / margin checker."),
        ("client_price_check_status", "Client Price Check Status", "Yes", "Status", "Yes", "Extension", "Pending / Checked / Issue."),
        ("basic_info_checked_by", "Basic Info Checked By", "No", "Text", "Yes", "Extension", "Basic order information checker."),
        ("basic_info_check_status", "Basic Info Check Status", "Yes", "Status", "Yes", "Extension", "Pending / Checked / Issue."),
        ("quality_requirement_checked_by", "Quality Requirement Checked By", "No", "Text", "Yes", "Extension", "Quality requirement checker, normally Harley."),
        ("quality_requirement_check_status", "Quality Requirement Check Status", "Yes", "Status", "Yes", "Extension", "Pending / Checked / Issue."),
        ("sample_requirement_confirmed", "Sample Requirement Confirmed", "Yes", "Yes/No", "Yes", "Extension", "Whether sample requirement is confirmed."),
        ("inspection_requirement_confirmed", "Inspection Requirement Confirmed", "Yes", "Yes/No", "Yes", "Extension", "Whether inspection requirement is confirmed."),
        ("target_delivery_date", "Target Delivery Date", "Recommended", "Date", "Controlled", "Order Details Mapping", "Target delivery date from Order Details."),
        ("po_sent_by", "PO Sent By", "No", "Text", "Yes", "Extension", "Person who sent PO to supplier."),
        ("po_sent_date", "PO Sent Date", "No", "Date", "Yes", "Extension", "PO sent date."),
        ("supplier_confirmation_received", "Supplier Confirmation Received", "No", "Yes/No", "Yes", "Extension", "Whether supplier confirmed PO."),
        ("release_type", "Release Type", "Yes", "Status", "Yes", "Extension", "Normal Release / Controlled Release."),
        ("release_status", "Release Status", "Yes", "Status", "Yes", "Extension", "Draft / Input Completed / Checking / Ready / Released / Hold."),
    ],
    "QP-04": [
        ("inspection_id", "Inspection ID", "Yes", "Text", "No", "Extension", "Unique inspection control record."),
        ("project_id", "Project ID", "Yes", "Text", "Controlled", "Existing system", "Linked project ID."),
        ("order_no", "Order No", "Yes", "Text", "Controlled", "Order Details", "Linked order number."),
        ("customer", "Customer", "Yes", "Text", "Controlled", "Existing system", "Customer code/name."),
        ("supplier_code", "Supplier Code", "Yes", "Text", "Controlled", "Supplier Details", "Supplier code."),
        ("supplier_name", "Supplier Name", "Yes", "Text", "No", "Supplier Details", "Supplier name."),
        ("product_status", "Product Status", "Yes", "Status", "Yes", "Extension", "New / Existing."),
        ("shipment_sequence_no", "Shipment Sequence No", "No", "Number", "Yes", "Extension", "Shipment sequence number."),
        ("inspection_required", "Inspection Required", "Yes", "Yes/No/By Case", "Yes", "Extension", "Whether inspection is required."),
        ("inspection_trigger_reason", "Inspection Trigger Reason", "No", "Text", "Yes", "Extension", "Reason inspection is required."),
        ("supplier_self_inspection_required", "Supplier Self-inspection Required", "Yes", "Yes/No", "Yes", "Extension", "Whether supplier self-inspection is required."),
        ("supplier_self_inspection_received", "Supplier Self-inspection Received", "Yes", "Yes/No", "Yes", "Extension", "Whether supplier self-inspection was received."),
        ("coa_required", "COA Required", "Yes", "Yes/No", "Yes", "Extension", "Whether Certificate of Analysis is required."),
        ("coa_received", "COA Received", "Yes", "Yes/No", "Yes", "Extension", "Whether COA was received."),
        ("coc_required", "COC Required", "Yes", "Yes/No", "Yes", "Extension", "Whether Certificate of Conformity is required."),
        ("coc_received", "COC Received", "Yes", "Yes/No", "Yes", "Extension", "Whether COC was received."),
        ("external_audit_required", "External Audit Required", "Yes", "Yes/No", "Yes", "Extension", "Whether third-party/external audit is required."),
        ("mark_inspection_required", "Mark Inspection Required", "Yes", "Yes/No", "Yes", "Extension", "Whether Mark inspection is required."),
        ("inspection_date", "Inspection Date", "No", "Date", "Yes", "Extension", "Inspection date."),
        ("inspection_result", "Inspection Result", "No", "Status", "Yes", "Extension", "Pass / Fail / Conditional Pass / Pending."),
        ("inspection_report_link", "Inspection Report Link", "No", "Link", "Yes", "Extension", "Inspection report link."),
        ("consecutive_pass_count", "Consecutive Pass Count", "No", "Number", "System / Controlled", "Extension / Derived", "Consecutive inspection pass count."),
        ("skip_shipment_balance", "Skip Shipment Balance", "No", "Number", "System / Controlled", "Extension / Derived", "Remaining shipments allowed to skip inspection."),
        ("deviation_exists", "Deviation Exists", "Yes", "Yes/No", "Yes", "Extension", "Whether deviation exists."),
        ("deviation_description", "Deviation Description", "No", "Text", "Yes", "Extension", "Deviation details."),
        ("customer_approval_required", "Customer Approval Required", "Yes", "Yes/No", "Yes", "Extension", "Whether customer approval is required before shipment."),
        ("customer_approval_status", "Customer Approval Status", "No", "Status", "Yes", "Extension", "Pending / Approved / Rejected."),
        ("shipment_release_status", "Shipment Release Status", "Yes", "Status", "Yes", "Extension", "Hold / Pending Inspection / Pending Customer / Ready / Released."),
        ("released_by", "Released By", "No", "Text", "Yes", "Extension", "Shipment release approver."),
        ("release_date", "Release Date", "No", "Date", "Yes", "Extension", "Shipment release date."),
    ],
    "QP-05": [
        ("complaint_id", "Complaint ID", "Yes", "Text", "No", "Extension", "Unique complaint record."),
        ("project_id", "Project ID", "Recommended", "Text", "Controlled", "Existing system", "Linked project ID."),
        ("order_no", "Order No", "Recommended", "Text", "Controlled", "Operation / Order Details", "Linked order number."),
        ("customer", "Customer", "Yes", "Text", "Controlled", "Existing system", "Customer code/name."),
        ("supplier_code", "Supplier Code", "Recommended", "Text", "Controlled", "Supplier Details", "Supplier code."),
        ("supplier_name", "Supplier Name", "Recommended", "Text", "No", "Supplier Details", "Supplier name."),
        ("issue_date", "Issue Date", "Yes", "Date", "Yes", "Extension", "Date issue was reported."),
        ("issue_type", "Issue Type", "Yes", "Status", "Yes", "Extension", "Product Quality / Inspection Failure / Packaging / Document / Shipment Damage / Other."),
        ("severity", "Severity", "Yes", "Status", "Yes", "Extension", "Critical / Major / Minor."),
        ("issue_description", "Issue Description", "Yes", "Text", "Yes", "Extension", "Issue description."),
        ("evidence_link", "Evidence Link", "No", "Link", "Yes", "Extension", "Evidence link."),
        ("immediate_correction", "Immediate Correction", "No", "Text", "Yes", "Extension", "Immediate containment/correction."),
        ("root_cause", "Root Cause", "No", "Text", "Yes", "Extension", "Root cause analysis."),
        ("corrective_action", "Corrective Action", "No", "Text", "Yes", "Extension", "Corrective action."),
        ("preventive_action", "Preventive Action", "No", "Text", "Yes", "Extension", "Preventive action."),
        ("supplier_response", "Supplier Response", "No", "Text", "Yes", "Extension", "Supplier reply."),
        ("customer_response", "Customer Response", "No", "Text", "Yes", "Extension", "Customer response."),
        ("responsible_person", "Responsible Person", "Yes", "Text", "Yes", "Extension", "Responsible person."),
        ("due_date", "Due Date", "No", "Date", "Yes", "Extension", "Follow-up due date."),
        ("closure_standard", "Closure Standard", "No", "Text", "Yes", "Extension", "Closure criteria."),
        ("closure_evidence", "Closure Evidence", "No", "Link/Text", "Yes", "Extension", "Closure evidence."),
        ("closed_by", "Closed By", "No", "Text", "Yes", "Extension", "Person who closed complaint."),
        ("close_date", "Close Date", "No", "Date", "Yes", "Extension", "Closure date."),
        ("supplier_history_updated", "Supplier History Updated", "Yes", "Yes/No", "Yes", "Extension", "Whether supplier history was updated."),
        ("complaint_status", "Complaint Status", "Yes", "Status", "Yes", "Extension", "Open / In Progress / Pending Supplier / Pending Customer / Closed."),
    ],
    "SV-01": [
        ("supplier_code", "Supplier Code", "Yes", "Text", "Controlled", "Supplier Details", "Supplier code."),
        ("supplier_name", "Supplier Name", "Yes", "Text", "Controlled", "Supplier Details", "Supplier name."),
        ("supplier_short_name", "Supplier Short Name", "No", "Text", "Yes", "Supplier Details", "Supplier short name."),
        ("main_products", "Main Products", "No", "Text", "Yes", "Supplier Details", "Main products."),
        ("main_process", "Main Process", "No", "Text", "Yes", "Supplier Details", "Main process."),
        ("country", "Country", "No", "Text", "Yes", "Supplier Details", "Country."),
        ("province", "Province", "No", "Text", "Yes", "Supplier Details", "Province."),
        ("city", "City", "No", "Text", "Yes", "Supplier Details", "City."),
        ("certificate_status", "Certificate Status", "No", "Text", "Yes", "Supplier Details", "Certificate status."),
        ("compliance_status", "Compliance Status", "No", "Status", "Yes", "Extension / Harley", "Compliance status."),
        ("quote_count", "Quote Count", "No", "Number", "No", "Derived", "Supplier quotation count."),
        ("order_count", "Order Count", "No", "Number", "No", "Derived", "Supplier order count."),
        ("sample_pass_rate", "Sample Pass Rate", "No", "Number", "No", "Derived", "Sample pass rate."),
        ("inspection_pass_rate", "Inspection Pass Rate", "No", "Number", "No", "Derived", "Inspection pass rate."),
        ("complaint_count", "Complaint Count", "No", "Number", "No", "Derived", "Complaint count."),
        ("open_complaint_count", "Open Complaint Count", "No", "Number", "No", "Derived", "Open complaint count."),
        ("last_quality_issue_date", "Last Quality Issue Date", "No", "Date", "No", "Derived", "Last quality issue date."),
        ("quality_risk_level", "Quality Risk Level", "No", "Status", "Yes", "Harley / Derived", "Supplier quality risk level."),
        ("cooperation_status", "Cooperation Status", "No", "Status", "Yes", "Harley / Derived", "Active / Watch / Hold / Archived."),
        ("supplier_quality_summary", "Supplier Quality Summary", "No", "Text", "Yes", "Harley / AI-assisted", "Supplier quality summary."),
        ("next_action", "Next Action", "No", "Text", "Yes", "Extension", "Next action."),
        ("owner", "Owner", "No", "Text", "Yes", "Extension", "Owner."),
    ],
    "HISTORY": [
        ("history_id", "History ID", "Yes", "Text", "No", "System", "Unique change history record."),
        ("process_code", "Process Code", "Yes", "Text", "No", "System", "Process code."),
        ("process_name", "Process Name", "Yes", "Text", "No", "System", "Process name."),
        ("version_from", "Version From", "No", "Text", "No", "System", "Previous version."),
        ("version_to", "Version To", "Yes", "Text", "Controlled", "System / Harley", "New version."),
        ("status", "Status", "Yes", "Status", "Controlled", "System", "Draft / Active / Archived."),
        ("change_date", "Change Date", "Yes", "DateTime", "No", "System", "Change date."),
        ("changed_by", "Changed By", "Yes", "Text", "No", "System", "Person who made change."),
        ("change_summary", "Change Summary", "Yes", "Text", "Yes", "Harley / AI-assisted", "Change summary."),
        ("change_detail", "Change Detail", "No", "Text", "Yes", "Harley / AI-assisted", "Change detail."),
        ("change_reason", "Change Reason", "No", "Text", "Yes", "Harley", "Reason for change."),
        ("change_impact_assessment", "Change Impact Assessment", "No", "Text", "Yes", "Harley / AI-assisted", "Potential impact on quality, delivery, cost or customer responsibility."),
        ("ai_generated_summary", "AI Generated Summary", "No", "Text", "Yes", "AI-assisted / Harley reviewed", "AI-assisted summary, generated only by button click."),
        ("harley_quality_approval", "Harley Quality Approval", "No", "Status", "Yes", "Extension", "Approved / Rejected / Pending."),
        ("harley_quality_approval_at", "Harley Approval Time", "No", "DateTime", "No", "System", "Harley approval timestamp."),
        ("harley_rejection_comment", "Harley Rejection Comment", "No", "Text", "Yes", "Extension", "Harley rejection comment."),
        ("maria_business_approval", "Maria Business Approval", "No", "Status", "Yes", "Extension", "Approved / Rejected / Pending."),
        ("maria_business_approval_at", "Maria Approval Time", "No", "DateTime", "No", "System", "Maria approval timestamp."),
        ("maria_rejection_comment", "Maria Rejection Comment", "No", "Text", "Yes", "Extension", "Maria rejection comment."),
        ("ehab_final_approval", "Ehab Final Approval", "No", "Status", "Yes", "Extension", "Approved / Rejected / Pending."),
        ("ehab_final_approval_at", "Ehab Approval Time", "No", "DateTime", "No", "System", "Ehab approval timestamp."),
        ("ehab_rejection_comment", "Ehab Rejection Comment", "No", "Text", "Yes", "Extension", "Ehab rejection comment."),
        ("effective_date", "Effective Date", "No", "Date", "Yes", "Harley", "Defaults to change date, can be future effective date."),
        ("archive_date", "Archive Date", "No", "Date", "No", "System", "Archive date."),
        ("remarks", "Remarks", "No", "Text", "Yes", "Extension", "Remarks."),
    ],
}

QUALITY_TEMPLATE_NAMES: dict[str, str] = {
    "QP-01 RFQ Requirement Control Template": "QP-01",
    "QP-02 Sample Control Template": "QP-02",
    "QP-03 Order Setup Control Template": "QP-03",
    "QP-04 Inspection & Shipment Release Template": "QP-04",
    "QP-05 Quality Complaint Closure Template": "QP-05",
    "SV-01 Supplier Management Template": "SV-01",
    "Process History Template": "HISTORY",
}


def get_process_definition(process_code: str) -> dict[str, Any]:
    return PROCESS_DEFINITIONS[process_code]


def list_process_definitions() -> list[dict[str, Any]]:
    return [PROCESS_DEFINITIONS[code] for code in PROCESS_ORDER]


def get_control_points(process_code: str) -> list[dict[str, str]]:
    return CONTROL_POINTS.get(process_code, [])


def available_quality_process_template_names() -> list[str]:
    return list(QUALITY_TEMPLATE_NAMES.keys())


def quality_process_template_name(process_code: str) -> str:
    for name, code in QUALITY_TEMPLATE_NAMES.items():
        if code == process_code:
            return name
    raise ValueError(f"No template name for process code {process_code}")


def _field_rows(process_code: str) -> list[dict[str, str]]:
    rows = []
    for field in QUALITY_TEMPLATE_FIELDS[process_code]:
        field_name, display_name, required, data_type, editable, source, description = field
        rows.append(
            {
                "field_name": field_name,
                "display_name": display_name,
                "required": required,
                "data_type": data_type,
                "editable_in_excel": editable,
                "source": source,
                "description": description,
            }
        )
    return rows


def _example_for_field(field_name: str) -> Any:
    examples: dict[str, Any] = {
        "rfq_id": "RFQ-202605-001",
        "sample_id": "SMP-202605-001",
        "order_setup_id": "OS-202605-001",
        "inspection_id": "INS-202605-001",
        "complaint_id": "CMP-202605-001",
        "history_id": "HIS-202605-001",
        "project_id": "SDG-26-001",
        "order_no": "EHS080526-1",
        "customer": "EHS",
        "supplier_code": "SUP-001",
        "supplier_name": "Example Supplier Co., Ltd.",
        "product_description": "Dolly wheel set",
        "rfq_working_file_link": "https://example.com/rfq-working-file",
        "customer_original_request_link": "https://example.com/customer-original-request",
        "sourcing_link": "https://example.com/sourcing-folder",
        "sampling_link": "https://example.com/sampling-folder",
        "design_file_link": "https://example.com/design-file",
        "quotation_to_client_link": "https://example.com/quotation-to-client",
        "original_requirement_notes": "URGENT project from EHS. Wheel diameter 40-50 mm, black preferred, loading capacity 30 kg, zinc coating, SST 72 hours.",
        "rfq_received_date": "2026-05-08",
        "rfq_received_by": "Maria",
        "drawing_received": "Yes",
        "specification_received": "Yes",
        "quantity_confirmed": "Yes",
        "target_price_received": "No",
        "delivery_requirement": "2026-06-30",
        "packaging_requirement": "See EHS packing detail",
        "testing_requirement": "Salt spray / coating thickness if required",
        "compliance_requirement": "REACH / PFAS / TSCA if applicable",
        "sample_required": "Yes",
        "inspection_required": "By Case",
        "harley_review_status": "Pending",
        "maria_review_status": "Pending",
        "rfq_gate_status": "Open",
        "current_owner": "Sandy",
        "due_date": "2026-05-12",
        "risk_level": "Medium",
        "sample_type": "Customer Sample",
        "customer_sample_qty": 2,
        "testing_sample_qty": 2,
        "retained_sample_qty": 1,
        "sample_plan_owner": "Sandy",
        "sample_status": "Open",
        "order_folder_link": "https://example.com/wecom-order-folder",
        "blank_order_excel_uploaded": "Yes",
        "order_input_by": "Sophia",
        "order_input_completed": "Yes",
        "quantity_check_status": "Pending",
        "supplier_price_check_status": "Pending",
        "client_price_check_status": "Pending",
        "quality_requirement_check_status": "Pending",
        "sample_requirement_confirmed": "Yes",
        "inspection_requirement_confirmed": "By Case",
        "target_delivery_date": "2026-06-30",
        "release_type": "Controlled Release",
        "release_status": "Checking",
        "product_status": "New",
        "supplier_self_inspection_required": "Yes",
        "supplier_self_inspection_received": "No",
        "coa_required": "Yes",
        "coa_received": "No",
        "coc_required": "Yes",
        "coc_received": "No",
        "external_audit_required": "Yes",
        "mark_inspection_required": "Yes",
        "inspection_result": "Pending",
        "deviation_exists": "No",
        "customer_approval_required": "Yes",
        "shipment_release_status": "Hold",
        "issue_date": "2026-05-08",
        "issue_type": "Product Quality",
        "severity": "Major",
        "responsible_person": "Harley",
        "supplier_history_updated": "No",
        "complaint_status": "Open",
        "process_code": "QP-01",
        "process_name": "RFQ Requirement Control",
        "version_to": "V1.0",
        "status": "Draft",
        "change_date": "2026-05-08",
        "changed_by": "Harley",
        "change_summary": "Initial process version created.",
        "change_impact_assessment": "No negative impact expected; improves control and traceability.",
        "effective_date": "2026-05-08",
        "owner": "Harley",
        "next_action": "Review supplier quality risk",
        "quality_risk_level": "Medium",
        "cooperation_status": "Active",
        "compliance_status": "Pending Review",
    }
    return examples.get(field_name, "")


def _style_workbook(writer: pd.ExcelWriter) -> None:
    wb = writer.book
    header_fill = PatternFill("solid", fgColor="111827")
    guide_fill = PatternFill("solid", fgColor="F3F4F6")
    required_fill = PatternFill("solid", fgColor="FEE2E2")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 10
            for cell in col_cells[:80]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 44))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 36))
        ws.auto_filter.ref = ws.dimensions
    if "Field Guide" in wb.sheetnames:
        ws = wb["Field Guide"]
        headers = [cell.value for cell in ws[1]]
        if "required" in headers:
            req_col = headers.index("required") + 1
            for row_idx in range(2, ws.max_row + 1):
                if str(ws.cell(row_idx, req_col).value or "").lower() == "yes":
                    for cell in ws[row_idx]:
                        cell.fill = required_fill
    if "Template" in wb.sheetnames and wb["Template"].max_row >= 2:
        for cell in wb["Template"][2]:
            cell.fill = guide_fill
            cell.font = Font(color="374151")



def _add_sheet_title(ws, title: str, subtitle: str | None = None) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=18, color="B91C1C")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells("A2:H2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="4B5563")
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 28


def _style_range(ws, cell_range: str, *, fill: str | None = None, font_color: str = "111827", bold: bool = False) -> None:
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color=font_color, bold=bold)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def _build_rfq_working_file_template(template_name: str) -> BytesIO:
    """RFQ template that keeps the old RFQ Working File strengths and adds the new RFQ Control Layer."""
    fields = _field_rows("QP-01")
    field_names = [row["field_name"] for row in fields]
    example = {field: _example_for_field(field) for field in field_names}

    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ Working File"
    _add_sheet_title(
        ws,
        "RFQ Working File + Control Layer",
        "Use this sheet as the familiar project working file: free notes and WeCom file links stay here. The control sections below add RFQ status, owner, next step and requirement tracking.",
    )
    widths = {"A": 22, "B": 34, "C": 22, "D": 22, "E": 20, "F": 20, "G": 20, "H": 24}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Link hub
    ws["A4"] = "Project File Links"
    ws["A4"].font = Font(bold=True, size=13, color="111827")
    link_rows = [
        ("Customer original request", "Paste customer RFQ/email/source link here", "customer_original_request_link"),
        ("Sourcing", "Paste sourcing/quotation folder link here", "sourcing_link"),
        ("Sampling", "Paste sampling folder link here", "sampling_link"),
        ("Design File", "Paste drawing/design file link here", "design_file_link"),
        ("Quotation to Client", "Paste client quotation file link here", "quotation_to_client_link"),
        ("System Project Link", "Optional system project link", "system_project_link"),
    ]
    ws.append([])
    start = 5
    ws.append(["Link Type", "Link / Description", "Mapped Field", "Owner", "Last Checked", "Remarks"])
    for item in link_rows:
        ws.append([item[0], item[1], item[2], "", "", ""])
    _style_range(ws, f"A{start}:F{start}", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(ws, f"A{start+1}:F{start+len(link_rows)}")

    # Basic info
    row = start + len(link_rows) + 3
    ws[f"A{row}"] = "Project Basic Information"
    ws[f"A{row}"].font = Font(bold=True, size=13, color="111827")
    row += 1
    headers = ["Client information", "Code", "QC/QA Level", "Exchange Rate", "Commission", "Date", "Operator"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row, idx).value = header
    ws.append(["", "EHS", "", "", "", "", "Sandy"])
    _style_range(ws, f"A{row}:G{row}", fill="D9EAD3", bold=True)
    _style_range(ws, f"A{row+1}:G{row+1}")

    # RFQ Control Summary
    row += 4
    ws[f"A{row}"] = "RFQ Control Summary"
    ws[f"A{row}"].font = Font(bold=True, size=13, color="B91C1C")
    row += 1
    summary = [
        ("RFQ ID", "RFQ-202605-001", "Harley/System"),
        ("Project ID", "SDG-26-001", "System / Sales Board"),
        ("RFQ Gate Status", "Open / Pending Information / Ready / Hold / Closed", "Harley"),
        ("Risk Level", "Low / Medium / High / Critical", "Harley + Maria / Ehab final"),
        ("Current Owner", "Sandy / Maria / Harley", "Process owner"),
        ("Next Step", "Confirm missing requirement / request supplier quote", "Owner"),
        ("Due Date", "YYYY-MM-DD", "Owner"),
    ]
    ws.append(["Control Field", "Value", "Responsible / Source", "Remarks"])
    for item in summary:
        ws.append([item[0], item[1], item[2], ""])
    _style_range(ws, f"A{row}:D{row}", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(ws, f"A{row+1}:D{row+len(summary)}")

    # Original notes
    row += len(summary) + 3
    ws[f"A{row}"] = "Original Requirement Notes"
    ws[f"A{row}"].font = Font(bold=True, size=13, color="B91C1C")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row+5, end_column=8)
    ws.cell(row, 1).value = "Paste original customer RFQ notes, product description, pictures reference, dimensions, quantities and special comments here. Keep this area flexible."
    ws.cell(row, 1).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor="FFF7ED")
    _style_range(ws, f"A{row}:H{row+5}")

    # Task/action log
    row += 8
    ws[f"A{row}"] = "Task / Action Log"
    ws[f"A{row}"].font = Font(bold=True, size=13, color="B91C1C")
    row += 1
    action_headers = ["No.", "Action", "Owner", "Due Date", "Status", "Result / Link", "Remarks"]
    for idx, header in enumerate(action_headers, start=1):
        ws.cell(row, idx).value = header
    for i in range(1, 6):
        ws.append([i, "", "", "", "Open / In Progress / Done", "", ""])
    _style_range(ws, f"A{row}:G{row}", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(ws, f"A{row+1}:G{row+5}")

    # Dedicated control summary sheet for structured RFQ gate management
    control = wb.create_sheet("RFQ Control Summary")
    _add_sheet_title(control, "RFQ Control Summary", "Standardised control area. Use this sheet to convert the free RFQ Working File into trackable system fields.")
    control_headers = ["Control Field", "Value", "Owner / Source", "System Field", "Remarks"]
    for idx, header in enumerate(control_headers, start=1):
        control.cell(4, idx).value = header
    control_rows = [
        ("RFQ ID", "RFQ-202605-001", "Harley / System", "rfq_id", "Unique RFQ control record"),
        ("Project ID", "SDG-26-001", "Sales Board", "project_id", "Linked project"),
        ("RFQ Working File Link", "https://example.com/rfq-working-file", "Operator", "rfq_working_file_link", "Main Excel / WeCom working file"),
        ("Customer", "EHS", "Sales Board", "customer", "Customer code/name"),
        ("Product Description", "1.5-2 inch casters", "RFQ Working File", "product_description", "Short RFQ item description"),
        ("RFQ Gate Status", "Open / Pending Information / Ready / Hold / Closed", "Harley", "rfq_gate_status", "Current gate status"),
        ("Missing Information", "", "Maria / Harley / Sandy", "missing_information", "Information required before supplier quotation or client quotation"),
        ("Risk Level", "Low / Medium / High / Critical", "Harley + Maria / Ehab final", "risk_level", "Overall risk level"),
        ("Current Owner", "Sandy", "Process owner", "current_owner", "Current follow-up owner"),
        ("Next Step", "Confirm missing information", "Owner", "next_step", "Next action"),
        ("Due Date", "YYYY-MM-DD", "Owner", "due_date", "Action due date"),
    ]
    for row_data in control_rows:
        control.append(row_data)
    _style_range(control, "A4:E4", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(control, f"A5:E{4+len(control_rows)}")
    for col in "ABCDE":
        control.column_dimensions[col].width = 30

    # Requirement checklist sheet
    req = wb.create_sheet("Requirement Checklist")
    _add_sheet_title(req, "RFQ Requirement Checklist", "Structured checklist added on top of the RFQ Working File. This is the main control layer for missing information and risk visibility.")
    req_headers = ["Requirement", "Confirmed?", "Source / Link", "Owner", "Risk / Remark"]
    for idx, header in enumerate(req_headers, start=1):
        req.cell(4, idx).value = header
    requirements = [
        "Drawing / design file", "Specification", "Quantity", "Target price", "Delivery requirement", "Packaging requirement", "Testing requirement", "Compliance requirement", "Sample requirement", "Inspection requirement", "Special customer requirement", "Missing information"
    ]
    for i, item in enumerate(requirements, start=5):
        req.cell(i, 1).value = item
        req.cell(i, 2).value = "Yes / No / By Case"
        req.cell(i, 4).value = "Maria / Harley / Sandy"
    _style_range(req, "A4:E4", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(req, f"A5:E{4+len(requirements)}")
    for col in "ABCDE":
        req.column_dimensions[col].width = 26

    # Risk review sheet
    risk = wb.create_sheet("Risk Review")
    _add_sheet_title(risk, "RFQ Risk Review", "Harley owns quality/compliance risk, Maria owns commercial/business risk, and Ehab confirms final risk level and action for major cases.")
    risk_headers = ["Risk Area", "Primary Owner", "Risk Level", "Risk Description", "Suggested Action", "Final Decision / Approval"]
    for idx, header in enumerate(risk_headers, start=1):
        risk.cell(4, idx).value = header
    risk_rows = [
        ("Quality risk", "Harley", "Low / Medium / High / Critical", "", "", ""),
        ("Regulatory / compliance risk", "Harley", "Low / Medium / High / Critical", "", "", ""),
        ("Customer contract / key-account requirement", "Harley", "Low / Medium / High / Critical", "", "", ""),
        ("New product / supplier / process risk", "Harley", "Low / Medium / High / Critical", "", "", ""),
        ("Specification deviation / non-conformance", "Harley", "Low / Medium / High / Critical", "", "", ""),
        ("Price / margin / hidden cost", "Maria", "Low / Medium / High / Critical", "", "", ""),
        ("Delivery / supply capability", "Maria", "Low / Medium / High / Critical", "", "", ""),
        ("Customer relationship / strategic customer", "Maria", "Low / Medium / High / Critical", "", "", ""),
        ("Unclear responsibility boundary", "Maria", "Low / Medium / High / Critical", "", "", ""),
        ("Final risk level and action", "Ehab", "Low / Medium / High / Critical", "", "", ""),
    ]
    for row_data in risk_rows:
        risk.append(row_data)
    _style_range(risk, "A4:F4", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(risk, f"A5:F{4+len(risk_rows)}")
    for col in "ABCDEF":
        risk.column_dimensions[col].width = 28

    # Import template, field guide and instructions
    template = wb.create_sheet("Template")
    for idx, name in enumerate(field_names, start=1):
        template.cell(1, idx).value = name
        template.cell(2, idx).value = example.get(name, "")
    _style_range(template, f"A1:{get_column_letter(len(field_names))}1", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(template, f"A2:{get_column_letter(len(field_names))}2", fill="F3F4F6")
    template.freeze_panes = "A2"
    template.auto_filter.ref = template.dimensions

    guide = wb.create_sheet("Field Guide")
    guide_headers = ["field_name", "display_name", "required", "data_type", "editable_in_excel", "source", "description"]
    for idx, header in enumerate(guide_headers, start=1):
        guide.cell(1, idx).value = header
    for r, row in enumerate(fields, start=2):
        for c, header in enumerate(guide_headers, start=1):
            guide.cell(r, c).value = row.get(header, "")
    _style_range(guide, f"A1:G1", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(guide, f"A2:G{len(fields)+1}")
    guide.freeze_panes = "A2"
    guide.auto_filter.ref = guide.dimensions

    instructions = wb.create_sheet("Instructions")
    instructions_data = [
        ("Positioning", "Old flow = RFQ Working File for free notes and file links. New flow = RFQ Control Layer for risk control, responsibility tracking and system records."),
        ("How to use", "Keep the familiar RFQ Working File. Add RFQ Control Summary, Requirement Checklist, Risk Review and Action Log."),
        ("Source of truth", "The system remains the source of truth after Harley imports or records key fields. Excel is the working file and backup."),
        ("Do not change", "Do not rename technical field names in the Template sheet."),
        ("QP-01 import note", "The Template sheet can be uploaded through Import Center > Extension Import > RFQ Requirement Control. Import is restricted to Harley."),
    ]
    instructions.append(["Item", "Note"])
    for row_data in instructions_data:
        instructions.append(row_data)
    _style_range(instructions, "A1:B1", fill="111827", font_color="FFFFFF", bold=True)
    _style_range(instructions, f"A2:B{len(instructions_data)+1}")
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 90

    for sheet in wb.worksheets:
        sheet.freeze_panes = sheet.freeze_panes or "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, col_cells in enumerate(sheet.columns, start=1):
            max_len = 10
            for cell in col_cells[:80]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 48))
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(sheet.column_dimensions[get_column_letter(col_idx)].width or 0, min(max_len + 2, 38))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def build_quality_process_template(template_name: str) -> BytesIO:
    if template_name not in QUALITY_TEMPLATE_NAMES:
        raise ValueError(f"Unsupported quality process template: {template_name}")
    process_code = QUALITY_TEMPLATE_NAMES[template_name]
    if process_code == "QP-01":
        return _build_rfq_working_file_template(template_name)
    definition = PROCESS_DEFINITIONS.get(process_code, {"process_code": "HISTORY", "process_name": "Process History"})
    fields = _field_rows(process_code)
    field_names = [row["field_name"] for row in fields]
    example = {field: _example_for_field(field) for field in field_names}
    template_df = pd.DataFrame([example], columns=field_names)
    guide_df = pd.DataFrame(fields)
    instructions_df = pd.DataFrame(
        [
            {"Item": "Template", "Note": template_name},
            {"Item": "Related Process", "Note": f"{definition.get('process_code', 'HISTORY')} - {definition.get('process_name', 'Process History')}"},
            {"Item": "Important", "Note": "Keep technical field names in row 1 unchanged."},
            {"Item": "Example row", "Note": "Row 2 is an example. Delete it before importing real data."},
            {"Item": "Source of truth", "Note": "The system remains the source of truth. Excel is a working template and backup format."},
            {"Item": "Current note", "Note": "QP-01 RFQ Requirement Control import is enabled. Other process imports will be connected later."},
        ]
    )
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        template_df.to_excel(writer, index=False, sheet_name="Template")
        guide_df.to_excel(writer, index=False, sheet_name="Field Guide")
        instructions_df.to_excel(writer, index=False, sheet_name="Instructions")
        _style_workbook(writer)
    output.seek(0)
    return output


def build_process_document_excel(process_code: str) -> BytesIO:
    definition = PROCESS_DEFINITIONS[process_code]
    def_df = pd.DataFrame([definition])
    control_df = pd.DataFrame(CONTROL_POINTS.get(process_code, []))
    field_df = pd.DataFrame(_field_rows(process_code))
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        def_df.to_excel(writer, index=False, sheet_name="Process Summary")
        control_df.to_excel(writer, index=False, sheet_name="Control Points")
        field_df.to_excel(writer, index=False, sheet_name="Template Fields")
        _style_workbook(writer)
    output.seek(0)
    return output


def build_history_document_excel() -> BytesIO:
    active_versions = pd.DataFrame(list_process_definitions())
    field_df = pd.DataFrame(_field_rows("HISTORY"))
    note_df = pd.DataFrame(
        [
            {"Item": "AI-assisted summary", "Note": "AI should be triggered only when Harley clicks Generate Summary. It must not run automatically on every change."},
            {"Item": "Impact assessment", "Note": "Change impact assessment should be generated by button and reviewed by Harley before saving."},
            {"Item": "Approval", "Note": "Approvals should include approval status, timestamp, and rejection comment fields."},
            {"Item": "Effective date", "Note": "Effective date can default to change date, but Harley can set a future effective date."},
        ]
    )
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        active_versions.to_excel(writer, index=False, sheet_name="Active Process Versions")
        field_df.to_excel(writer, index=False, sheet_name="History Fields")
        note_df.to_excel(writer, index=False, sheet_name="Design Notes")
        _style_workbook(writer)
    output.seek(0)
    return output


def quality_template_file_name(template_name: str) -> str:
    safe = (
        template_name.lower()
        .replace(" / ", "_")
        .replace("&", "and")
        .replace("(", "")
        .replace(")", "")
        .replace(" ", "_")
        .replace("__", "_")
    )
    return f"{safe}.xlsx"


def process_document_file_name(process_code: str) -> str:
    definition = PROCESS_DEFINITIONS[process_code]
    safe = definition["short_name"].lower().replace(" / ", "_").replace("&", "and").replace(" ", "_")
    return f"{process_code.lower()}_{safe}_process_document.xlsx"
