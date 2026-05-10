from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, Sequence

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.template_service import build_import_template, template_file_name
from services.process_management_service import build_quality_process_template, quality_template_file_name

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join(ch for ch in str(name or "Records") if ch not in r"[]:*?/\\")
    return (cleaned or "Records")[:31]


def _safe_file_name(name: str) -> str:
    cleaned = str(name or "export").lower().strip()
    for old, new in {
        " / ": "_",
        "/": "_",
        "\\": "_",
        "&": "and",
        " ": "_",
        ":": "_",
        "·": "_",
        "__": "_",
    }.items():
        cleaned = cleaned.replace(old, new)
    cleaned = "".join(ch for ch in cleaned if ch.isalnum() or ch in {"_", "-", "."})
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_") or "export"


def _style_workbook(writer: pd.ExcelWriter) -> None:
    wb = writer.book
    header_fill = PatternFill("solid", fgColor="111827")
    sub_fill = PatternFill("solid", fgColor="F3F4F6")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        if ws.max_row >= 2 and ws.title.lower().startswith("export"):
            for cell in ws[2]:
                cell.fill = sub_fill
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 10
            for cell in col_cells[:120]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 46))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 38))


def records_to_excel(
    rows: Iterable[dict] | pd.DataFrame | None,
    *,
    sheet_name: str = "Records",
    title: str | None = None,
    note: str | None = None,
) -> bytes:
    """Build a lightweight Excel export for any board view without writing to the database."""
    if isinstance(rows, pd.DataFrame):
        df = rows.copy()
    else:
        df = pd.DataFrame(list(rows or []))
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=_safe_sheet_name(sheet_name))
        info = pd.DataFrame(
            [
                {"Item": "Title", "Value": title or sheet_name},
                {"Item": "Exported At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                {"Item": "Row Count", "Value": len(df)},
                {"Item": "Note", "Value": note or "Export is read-only. System records remain the source of truth."},
            ]
        )
        info.to_excel(writer, index=False, sheet_name="Export Info")
        _style_workbook(writer)
    output.seek(0)
    return output.getvalue()


def build_template_download(template_name: str) -> tuple[bytes, str]:
    """Use the shared template source for Import Center and board pages."""
    if template_name.startswith(("QP-", "SV-")) or template_name == "Process History Template":
        return build_quality_process_template(template_name).getvalue(), quality_template_file_name(template_name)
    return build_import_template(template_name).getvalue(), template_file_name(template_name)


def render_standard_export_panel(
    *,
    board_name: str,
    current_rows: Iterable[dict] | pd.DataFrame | None,
    filtered_rows: Iterable[dict] | pd.DataFrame | None = None,
    template_names: Sequence[str] | None = None,
    key_prefix: str,
) -> None:
    """Render a consistent export area for business boards.

    - Export is available to all logged-in users.
    - Import stays restricted in Import Center.
    - Templates are generated from the same central services used by Import Center.
    """
    template_names = list(template_names or [])
    filtered_rows = current_rows if filtered_rows is None else filtered_rows
    current_count = len(current_rows) if not isinstance(current_rows, pd.DataFrame) else len(current_rows.index)
    filtered_count = len(filtered_rows) if not isinstance(filtered_rows, pd.DataFrame) else len(filtered_rows.index)

    with st.expander("Export & templates", expanded=False):
        st.caption(
            "Export is available to all logged-in users. Import remains restricted to Harley / authorised emails in Import Center. "
            "System records are the source of truth; Excel is a working copy and backup format."
        )
        export_col, filtered_col, template_col = st.columns([1, 1, 1.25])
        with export_col:
            st.download_button(
                "Export current view",
                data=records_to_excel(
                    current_rows,
                    sheet_name="Current View",
                    title=f"{board_name} - Current View",
                    note=f"Current view export from {board_name}.",
                ),
                file_name=f"{_safe_file_name(board_name)}_current_view.xlsx",
                mime=EXCEL_MIME,
                use_container_width=True,
                key=f"{key_prefix}_export_current",
                help=f"Exports the current {board_name} dataset ({current_count} row(s)).",
            )
        with filtered_col:
            st.download_button(
                "Export filtered records",
                data=records_to_excel(
                    filtered_rows,
                    sheet_name="Filtered Records",
                    title=f"{board_name} - Filtered Records",
                    note=f"Filtered record export from {board_name}.",
                ),
                file_name=f"{_safe_file_name(board_name)}_filtered_records.xlsx",
                mime=EXCEL_MIME,
                use_container_width=True,
                key=f"{key_prefix}_export_filtered",
                help=f"Exports filtered {board_name} records ({filtered_count} row(s)).",
            )
        with template_col:
            if not template_names:
                st.info("No import/update template is configured for this board yet.")
            else:
                for idx, template_name in enumerate(template_names):
                    try:
                        template_bytes, file_name = build_template_download(template_name)
                    except Exception as exc:  # keep page usable if a template definition is missing
                        st.warning(f"Template unavailable: {template_name} ({type(exc).__name__})")
                        continue
                    st.download_button(
                        f"Download {template_name}",
                        data=template_bytes,
                        file_name=file_name,
                        mime=EXCEL_MIME,
                        use_container_width=True,
                        key=f"{key_prefix}_template_{idx}",
                    )
