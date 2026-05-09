from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.import_service import ALL_IMPORT_FIELDS, IMPORT_FIELDS, OPTIONAL_FIELDS
from services.upgrade_service import MODULES, import_field_names, import_required_fields


CORE_TEMPLATE_LABELS: dict[str, dict[str, str]] = {
    "Sales": {
        "project_id": "Project ID",
        "project_name": "Project Name",
        "client_code": "Client Code",
        "category": "Category",
        "priority": "Priority",
        "reference_link": "Reference Link",
    },
    "Operation": {
        "project_id": "Project ID",
        "client_code": "Client Code",
        "order_no": "Order No",
        "reference_link": "Reference Link",
    },
}

CORE_TEMPLATE_DESCRIPTIONS: dict[str, dict[str, str]] = {
    "Sales": {
        "project_id": "Required. Unique Sales project ID, for example SDG-26-001.",
        "project_name": "Required. Project name or short description.",
        "client_code": "Required. Client/customer code.",
        "category": "Optional. Project category.",
        "priority": "Optional. Priority level.",
        "reference_link": "Optional. WeCom folder, document, or other project reference link.",
    },
    "Operation": {
        "project_id": "Required. Linked Sales project ID. If not yet available, use the best known project/order reference.",
        "client_code": "Required. Client/customer code.",
        "order_no": "Required. Unique operation order number.",
        "reference_link": "Optional. WeCom order folder, document, or other order reference link.",
    },
}

CORE_TEMPLATE_EXAMPLES: dict[str, dict[str, Any]] = {
    "Sales": {
        "project_id": "SDG-26-001",
        "project_name": "Example product RFQ",
        "client_code": "EHS",
        "category": "Hardware",
        "priority": "High",
        "reference_link": "https://example.com/wecom-folder-link",
    },
    "Operation": {
        "project_id": "SDG-26-001",
        "client_code": "EHS",
        "order_no": "EHS080526-1",
        "reference_link": "https://example.com/wecom-order-folder-link",
    },
}

TEMPLATE_MODULE_MAP = {
    "Sales Import Template": ("core", "Sales"),
    "Operation Import Template": ("core", "Operation"),
    "Supplier Details Template": ("extension", "Supplier Details"),
    "Price Comparison Template": ("extension", "Supplier Price Comparison"),
    "Order Details Template": ("extension", "Order Details"),
    "Sample Tracking Template": ("extension", "Sample Tracking"),
}


def available_template_names() -> list[str]:
    return list(TEMPLATE_MODULE_MAP.keys())


def _extension_field_meta(module_name: str) -> dict[str, Any]:
    fields = {field.name: field for field in MODULES[module_name].fields}
    import_names = import_field_names(module_name)
    required = set(import_required_fields(module_name))
    return {
        "fields": import_names,
        "required": required,
        "labels": {name: fields[name].display for name in import_names if name in fields},
        "descriptions": {name: fields[name].description for name in import_names if name in fields},
        "numeric": {name for name in import_names if getattr(fields.get(name), "numeric", False)},
        "boolean": {name for name in import_names if getattr(fields.get(name), "boolean", False)},
    }


def _core_field_meta(import_type: str) -> dict[str, Any]:
    fields = ALL_IMPORT_FIELDS[import_type]
    return {
        "fields": fields,
        "required": set(IMPORT_FIELDS[import_type]),
        "labels": CORE_TEMPLATE_LABELS.get(import_type, {}),
        "descriptions": CORE_TEMPLATE_DESCRIPTIONS.get(import_type, {}),
        "numeric": set(),
        "boolean": set(),
    }


def _example_value(field_name: str, *, module_name: str | None = None, import_type: str | None = None) -> Any:
    if import_type and field_name in CORE_TEMPLATE_EXAMPLES.get(import_type, {}):
        return CORE_TEMPLATE_EXAMPLES[import_type][field_name]

    sample_values = {
        "project_id": "SDG-26-001",
        "rfq_item_ref": "RFQ-001",
        "order_no": "EHS080526-1",
        "order_detail_id": "",
        "sample_id": "",
        "supplier_id": "",
        "supplier_code": "SUP-001",
        "supplier_name": "Example Supplier Co., Ltd.",
        "supplier_short_name": "Example Supplier",
        "client_code": "EHS",
        "order_item_code": "ITEM-001",
        "supplier_quote_id": "",
        "supplier_quote_id": "",
        "supplier_quote_id": "",
        "supplier_unit_cost": 1.25,
        "client_unit_price": 1.85,
        "currency": "USD",
        "order_qty": 1000,
        "unit": "set",
        "order_date": "2026-05-08",
        "target_delivery_date": "2026-06-30",
        "actual_delivery_date": "",
        "shipment_date": "",
        "quote_date": "2026-05-08",
        "sample_request_date": "2026-05-08",
        "target_sample_date": "2026-05-18",
        "sample_status": "In Progress",
        "sample_type": "Initial",
        "sample_purpose": "Client Approval / Testing / Reference",
        "testing_required": "Yes",
        "revision_required": "No",
        "next_sample_round_needed": "No",
        "sample_folder_link": "https://example.com/sample-folder-link",
        "payment_status": "Pending",
        "production_status": "Not Started",
        "inspection_status": "Pending",
        "packing_status": "Pending",
        "shipment_status": "Pending",
        "main_issue": "",
        "next_step": "Confirm sample plan",
        "next_step_owner": "Harley",
        "remarks": "Example row. Delete before import.",
        "reference_link": "https://example.com/wecom-folder-link",
        "supplier_name": "Example Supplier Co., Ltd.",
        "company_type": "Factory",
        "country": "China",
        "province": "Guangdong",
        "city": "Shenzhen",
        "primary_contact_name": "Contact Name",
        "primary_contact_mobile": "",
        "primary_contact_email": "",
        "main_products": "Hardware fittings",
        "quality_risk": "Medium",
        "commercial_risk": "Medium",
        "recommended_supplier": "Yes",
        "selected_supplier": "No",
        "quotation_quality": "Complete",
        "quotation_risk": "Medium",
        "comparison_status": "In Progress",
    }
    return sample_values.get(field_name, "")


def _style_workbook(writer: pd.ExcelWriter, template_sheet: str = "Template") -> None:
    wb = writer.book
    header_fill = PatternFill("solid", fgColor="111827")
    guide_fill = PatternFill("solid", fgColor="F3F4F6")
    required_fill = PatternFill("solid", fgColor="FEE2E2")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 10
            for cell in col_cells[:80]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 42))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 34))

    if template_sheet in wb.sheetnames:
        ws = wb[template_sheet]
        ws.freeze_panes = "A2"
        # Make example row visually different so it is not mistaken as required data.
        if ws.max_row >= 2:
            for cell in ws[2]:
                cell.fill = guide_fill
                cell.font = Font(color="374151")
        # Add filter to template header.
        ws.auto_filter.ref = ws.dimensions

    if "Field Guide" in wb.sheetnames:
        ws = wb["Field Guide"]
        headers = [cell.value for cell in ws[1]]
        if "required" in headers:
            req_col = headers.index("required") + 1
            for row_idx in range(2, ws.max_row + 1):
                if str(ws.cell(row_idx, req_col).value or "").lower() == "yes":
                    for cell in ws[row_idx]:
                        cell.fill = required_fill


def build_import_template(template_name: str) -> BytesIO:
    if template_name not in TEMPLATE_MODULE_MAP:
        raise ValueError(f"Unsupported template: {template_name}")

    kind, name = TEMPLATE_MODULE_MAP[template_name]
    if kind == "core":
        meta = _core_field_meta(name)
        fields = meta["fields"]
        example = {field: _example_value(field, import_type=name) for field in fields}
        title = f"{name} Import Template"
        description = "Core Sales / Operation import template. Import is restricted to authorised users."
    else:
        meta = _extension_field_meta(name)
        fields = meta["fields"]
        example = {field: _example_value(field, module_name=name) for field in fields}
        title = f"{name} Import Template"
        description = "Extension import template. It writes to extension tables and does not change core Sales / Operation logic."

    required = meta["required"]
    labels = meta["labels"]
    descriptions = meta["descriptions"]
    numeric = meta["numeric"]
    boolean = meta["boolean"]

    template_df = pd.DataFrame([example], columns=fields)
    guide_df = pd.DataFrame(
        [
            {
                "field_name": field,
                "display_name": labels.get(field, field),
                "required": "Yes" if field in required else "No",
                "data_type": "Boolean Yes/No" if field in boolean else ("Number" if field in numeric else "Text / Date / Link"),
                "description": descriptions.get(field, ""),
            }
            for field in fields
        ]
    )
    notes_df = pd.DataFrame(
        [
            {"Item": "Template", "Note": title},
            {"Item": "Usage", "Note": description},
            {"Item": "Important", "Note": "Keep the technical field names in row 1 unchanged. The system import mapping can match these headers directly."},
            {"Item": "Example row", "Note": "Row 2 is only an example. Delete it before importing real data."},
            {"Item": "Source of truth", "Note": "The system remains the source of truth. Excel is a working template and backup format."},
        ]
    )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        template_df.to_excel(writer, index=False, sheet_name="Template")
        guide_df.to_excel(writer, index=False, sheet_name="Field Guide")
        notes_df.to_excel(writer, index=False, sheet_name="Instructions")
        _style_workbook(writer)
    output.seek(0)
    return output


def template_file_name(template_name: str) -> str:
    safe = template_name.lower().replace(" / ", "_").replace(" ", "_").replace("__", "_")
    return f"{safe}.xlsx"
